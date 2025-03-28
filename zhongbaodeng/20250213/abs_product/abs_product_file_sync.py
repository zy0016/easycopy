# coding=utf-8
# !/usr/bin/python
#-------------------------------------------------------------------------------
# Name:        abs_product_file_sync
# Purpose:     同步资产资产计划abs提交受理处理后的附件信息
#
# Author:      lihuaizhong
#
# Created:     10-11-2020
# Copyright:   (c) lihuaizhong 2020
# Licence:     <your licence>
#-------------------------------------------------------------------------------
# 获取估值需要的产品信息列表和附件

import logging
import os
import os.path
import socket
import sys
import time
from datetime import datetime
import paramiko
from openpyxl import Workbook
from abs_product.func import savelog, get_date_info_map, excel_data, get_increase_str,read_config,excel_data_month,switch_reg_date,get_default_value,list_dates_between,str_is_float,get_truncate_date
from abs_product.office_func import create_excel_file,create_reg_month_excel_file


from docxtpl import DocxTemplate
try:
    import cx_Oracle
except ImportError as e:
    print("导入cx_Oracle模块失败,需要安装cx_Oracle\r\n{}",str(e))
    sys.exit(-1)

logging.basicConfig()
configmap = read_config()
os.environ["ORACLE_SID"] = str(configmap["oracle_sid"])
os.environ["ORACLE_BASE"] = str(configmap["oracle_base"])
os.environ["ORACLE_HOME"] = str(configmap["oracle_home"])
os.environ["LD_LIBRARY_PATH"] = str(configmap["ld_library_path"])

# 附件catalog
catalog = 'TFUNDINFO'

class SSHSession(object):

    def __init__(self,fileSrcHost,fileSrcPort,fileSrcUserName,fileSrcPassword,dbip,dbport,dbSID,dbUserName,dbPassword,key_file=None):

        self.fileSrcHost = fileSrcHost
        self.fileSrcPort = fileSrcPort
        self.fileSrcUserName = fileSrcUserName
        self.fileSrcPassword = fileSrcPassword

        self.dbip = dbip
        self.dbport = dbport
        self.dbSID = dbSID
        self.dbusername = dbUserName
        self.dbpassword = dbPassword

        self.fileSrcSftp = self.getFtp(self.fileSrcHost,self.fileSrcPort,self.fileSrcUserName,self.fileSrcPassword)
        self.cursor = self.getDbCursor(self.dbip,self.dbport,self.dbSID,self.dbusername,self.dbpassword)

    def getDbCursor(self,dbIp,dbPort,dbSID,dbUserName,dbPassword):
        try:
            dsn_tns = cx_Oracle.makedsn(dbIp, dbPort, service_name=dbSID)
            self.db = cx_Oracle.connect(dbUserName, dbPassword, dsn_tns)
            cursor = self.db.cursor()
            print("connet %s:%s/%s database successful!" % (dbIp,dbPort,dbSID))
            return cursor
        except cx_Oracle.Error as e:
            print("init connet database failed,exit!" + str(e))
            time.sleep(5)
            sys.exit(-1)

    def getFtp ( self ,host,port,userName,userPassword):
        print("connect fileserver %s:%s" % (host,port))
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(30 * 60 * 1000)
        sock.connect((host, port))
        t = paramiko.Transport(sock)
        t.start_client()
        # supposed to check for key in keys, but I don't much care right now to find the right notation
        if userPassword is not None:
            t.auth_password(userName, userPassword, fallback=False)
        else:
            raise Exception('Must supply either key_file or password')
        print("connect fileserver %s:%s Successful!!" % (host, port))
        return paramiko.SFTPClient.from_transport(t)

    def reconnectDB(self):
        try:
            print("reconnet %s:%s/%s database " % (self.dbip,self.dbport,self.dbSID))
            self.cursor = self.getDbCursor(self.dbip,self.dbport,self.dbSID,self.dbUserName,self.dbPassword)
            print("reconnet %s:%s/%s database successful!" % (self.dbip,self.dbport,self.dbSID))
        except cx_Oracle.Error as e:
            print("reconnet database failed,exit!" + str(e))
            time.sleep(5)
            sys.exit(-1)

    def disconnect(self):
        try:
            self.fileSrcSftp.close()
            self.db.close()
            print("Connection terminated.")
        except cx_Oracle.DatabaseError as e:
            print("close db  except." + str(e))

    # 获取产品代码：为产品注册流程的分配受理节点提交受理通过后的产品代码
    def getFundcodes(self,templetId,startDate,endDate,localpath):
        fundcodes = set()
        sql = "select distinct f.c_fundcode,f.c_fundname from tfundinfo f where f.l_pdttempletid ='"+templetId+"' and f.c_fundcode in (select j.BIZ_FLOW_UUID_  from jbpm4_ext_htask_ j  where j.PROC_NAME_='产品注册流程'  and j.ACTIVITY_NAME_='分配受理' and j.TASKOUTCOMEACTION='提交受理' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>='"+startDate+"' and to_char(j.END_,'yyyymmdd')<='"+endDate+"' \
               and  j.BIZ_FLOW_UUID_ not in (select j.BIZ_FLOW_UUID_  from jbpm4_ext_htask_ j where j.PROC_NAME_='产品注册流程'  and j.ACTIVITY_NAME_='中保登受理' and to_char(j.END_,'yyyymmdd')>='"+startDate+"' and to_char(j.END_,'yyyymmdd')<='"+endDate+"')) "
        print('getFundcode.startDate = '+ startDate+',endDate='+endDate+',templetName='+ self.getTempleName(templetId),'sql='+sql)
        try:
            dataResult = self.cursor.execute(sql)
            data = dataResult.fetchall()
            for r in range(len(data)):
                fundcode = data[r][0]
                tafundcode = data[r][1]
                fundcodes.add(fundcode+','+tafundcode)
            return fundcodes

        except Exception as e:
            print('获取templetId=' + templetId + '类型的产品信息异常'+str(e))
        return fundcodes
    #获取组合类统计报告
    def getCombProductReport(self,startDate,endDate,fileLocalTemplateFile,localpath):
        sql = "select substr("+endDate+",1,4) reportYear,substr("+endDate+",5,2) reportMonth,to_number(substr("+endDate+",7,2)) reportDay,\
         a.mgrorgnoCount,a.combproductCount,a.sucessCount,b.per24hCount+b1.per24hCount1,c.nonAcceptCount,a.backCount,d.setreportCount,\
         a.fixedincomeCount,a.fixedIncomeRatio,a.mixCount,a.rightsCount, \
         a.mixRatio,a.rightsRatio,a.structuredCount,a.nonStructuredCount,e.istoinsureassetCount, \
         e.isincludepersonalCount,e.isfirstCount,e.ismonCount,e.ismoneyflagCount,F.isAutoCacelCount,F.isApplyCacelCount,F.cacelCount,e.isfofCount from ( \
   /*产品管理人,报送产品总量*/  \
      select count(distinct b.c_mgrorgno) mgrorgnoCount,count(distinct f.c_fundcode) combproductCount,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',f.c_fundcode,'机构复核通过',f.c_fundcode)) sucessCount,\
          count(distinct decode(j.TASKOUTCOMEACTION,'不予登记',f.c_fundcode)) backCount,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'0',f.c_fundcode))) fixedincomeCount,\
          round(decode(count(distinct f.c_fundcode),0,0,count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'0',f.c_fundcode)))/count(distinct decode(j.TASKOUTCOMEACTION, '准予登记', f.c_fundcode)))*100, 2)||'%' fixedIncomeRatio,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'9',f.c_fundcode ))) mixCount,\
          round(decode(count(distinct f.c_fundcode),0,0,count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'9',f.c_fundcode )))/count(distinct decode(j.TASKOUTCOMEACTION, '准予登记', f.c_fundcode)))*100, 2)||'%' mixRatio,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'1',f.c_fundcode ))) rightsCount,\
          round(decode(count(distinct f.c_fundcode),0,0,count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'1',f.c_fundcode )))/count(distinct decode(j.TASKOUTCOMEACTION, '准予登记', f.c_fundcode)))*100, 2)||'%' rightsRatio,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(f.l_pdttempletid,'10',f.c_fundcode))) nonStructuredCount,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(f.l_pdttempletid,'17',f.c_fundcode))) structuredCount\
    from tfundinfo f ,tpdtbasicinfo  b,(select j.BIZ_FLOW_UUID_ c_fundcode,j.TASKOUTCOMEACTION  from jbpm4_ext_htask_ j  where j.PROC_NAME_ in('发行前登记流程','发行前快速登记流程') and j.ACTIVITY_NAME_ in ( '机构复核','中保登审批','中保登复核')\
   and j.OUTCOME_ in ('pass','oppose') and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+") j where f.l_pdttempletid in ('10','17') \
   and f.c_fundcode=b.c_fundcode and  f.c_fundcode=j.c_fundcode)a,(\
   /*24小时内予以登记的产品*/ \
   select count(distinct a. BIZ_FLOW_UUID_) per24hCount from (\
   select j.BIZ_FLOW_UUID_,max(j.END_) END_  from jbpm4_ext_htask_ j  where j.PROC_NAME_='发行前登记流程' and j.ACTIVITY_NAME_ in ( '中保登审批','中保登复核')\
   and j.TASKOUTCOMEACTION='准予登记' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+" group by j.BIZ_FLOW_UUID_)a,(\
    select j.BIZ_FLOW_UUID_,min(j.CREATE_ )CREATE_ from jbpm4_ext_htask_ j  where j.PROC_NAME_='发行前登记流程' and j.ACTIVITY_NAME_='分配受理'\
   and j.TASKOUTCOMEACTION='提交受理' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+"  group by j.BIZ_FLOW_UUID_  )b\
   where a.BIZ_FLOW_UUID_=b.BIZ_FLOW_UUID_ and TO_NUMBER((cast(a.END_ AS DATE)- cast(b.CREATE_ AS DATE))*24)<=24)b,(\
   select count(distinct decode(j.TASKOUTCOMEACTION,'机构复核通过',f.c_fundcode)) per24hCount1\
    from tfundinfo f ,tpdtbasicinfo  b,(select j.BIZ_FLOW_UUID_ c_fundcode,j.TASKOUTCOMEACTION  from jbpm4_ext_htask_ j  where j.PROC_NAME_ in('发行前快速登记流程') and j.ACTIVITY_NAME_ in ( '机构复核')\
   and j.OUTCOME_ in ('pass') and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+") j where f.l_pdttempletid in ('10','17') \
   and f.c_fundcode=b.c_fundcode and  f.c_fundcode=j.c_fundcode)b1,(\
   /*不予受理的产品*/ \
   select count(distinct BIZ_FLOW_UUID_) nonAcceptCount  from jbpm4_ext_htask_ j  where j.PROC_NAME_='发行前登记流程' and j.ACTIVITY_NAME_='不予受理复核'\
   and j.TASKOUTCOMEACTION='复核通过' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+" )c,(\
   /*产品设立报告*/ \
   select count(distinct f.c_fundcode)  setreportCount from  tfundinfo f ,tblobstorage bs where f.l_pdttempletid in ('10','17') and  f.c_fundcode in (\
   select j.BIZ_FLOW_UUID_  from jbpm4_ext_htask_ j  where j.PROC_NAME_ in ( '产品设立报告流程','设立报告提交流程') and j.ACTIVITY_NAME_ in  ('产品设立报告审核', '平台接收')\
   and j.TASKOUTCOMEACTION='接收' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+" )\
   and f.c_fundcode=bs.c_subkeyid and bs.c_memo='产品设立报告')d,(\
   /*MOM产品*/ \
    select\
           count(decode(b.C_ISTOINSUREASSET,'0',f.c_fundcode,null  )) istoinsureassetCount ,/*面向非保险资金募集只数*/ \
           count(decode(b.C_ISINCLUDEPERSONAL,'1',f.c_fundcode,null  )) isincludepersonalCount,/*面向自然人投资者募集的产品只数*/ \
           count(decode(b.C_ISFIRSTPRODUCT,'1',f.c_fundcode,null  )) isfirstCount,/*首单产品只数*/ \
           count(decode(b.C_ISMMTYPE,'1',f.c_fundcode,null  )) ismonCount,/*MOM产品只数*/ \
           count(decode(b.C_ISFFTYPE,'1',f.c_fundcode,null  )) isfofCount,/*FOF产品只数*/ \
           count(decode(b.C_ISMONEYFLAG,'1',f.c_fundcode,null )) ismoneyflagCount/*货币市场类产品只数*/ \
    from tfundinfo f ,TPDTASSORTEDPRODUCTINFO  b where f.l_pdttempletid in ('10','17') and  f.c_fundcode in (\
   select j.BIZ_FLOW_UUID_  from jbpm4_ext_htask_ j  where j.PROC_NAME_ in('发行前登记流程','发行前快速登记流程') and j.ACTIVITY_NAME_ in ( '机构复核','中保登审批','中保登复核')\
   and j.TASKOUTCOMEACTION in('准予登记','机构复核通过') and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+" )\
   and f.c_fundcode=b.c_fundcode)e,( SELECT \
          COUNT(DECODE(C.C_CANCELTYPE,'0',A.C_FUNDCODE)) isAutoCacelCount, \
          COUNT(DECODE(C.C_CANCELTYPE,'1',A.C_FUNDCODE)) isApplyCacelCount, \
          COUNT(A.C_FUNDCODE) cacelCount \
   FROM TFUNDINFO A JOIN TPDTBASICINFO B ON A.C_FUNDCODE = B.C_FUNDCODE JOIN TPDTEXTENDINFO C ON A.C_FUNDCODE = C.C_FUNDCODE \
   WHERE  A.C_CLIENTAUDITFLAG = 48 AND TO_CHAR(B.D_CANCELDATE,'yyyymmdd') <= "+endDate+") F"
        print('getCombProductReport sql='+sql)
        try:
            if not os.path.isfile(fileLocalTemplateFile):
                print('模板文件不存在:'+fileLocalTemplateFile)
                exit()
            tpl = DocxTemplate(fileLocalTemplateFile)
            dataResult = self.cursor.execute(sql)
            data = dataResult.fetchall()
            for r in range(len(data)):

                context = {'year': data[r][0],'month': data[r][1],'day': data[r][2],'mgrorgnoCount': data[r][3],
                           'combproductCount': data[r][4],'sucessCount': data[r][5],'per24hCount': data[r][6],'nonAcceptCount': data[r][7],
                           'backCount': data[r][8],'setreportCount': data[r][9],'fixedincomeCount': data[r][10],'fixedIncomeRatio': data[r][11],
                           'mixCount': data[r][12],'rightsCount': data[r][13],'mixRatio': data[r][14],'rightsRatio': data[r][15],
                           'structuredCount': data[r][16],'nonStructuredCount': data[r][17],'istoinsureassetCount': data[r][18],'isincludepersonalCount': data[r][19],
                           'isfirstCount': data[r][20],'ismonCount': data[r][21],'ismoneyflagCount': data[r][22],'isAutoCacelCount': data[r][23],'isApplyCacelCount': data[r][24],'cacelCount': data[r][25],
                           'isfofCount': data[r][26]
                          }
            tpl.render(context)
            #tmpreportpath = localpath +os.path.sep +'combreport'+ os.path.sep
            tmpreportpath = localpath +os.path.sep + endDate + os.path.sep + 'annex'+ os.path.sep
            self.mkdir_ifnotexists(tmpreportpath)
            tpl.save(tmpreportpath+'Comb_product_Report_'+endDate+'.docx')
            print('生成'+tmpreportpath+' 成功')
        except Exception as e:
            print('获取组合类型的产品信息异常'+str(e))
    # 获取融资主体/担保主体编码
    def getMainCode(self,name):
        if name is None:
            return ''
        if '' == name.strip():
            return name.strip()
        # sql = "  select C_ORGCODE from TPDTMANDATOR where c_name = '"+name+"'"
        sql = " select C_ORGCODE from TPDTMANDATOR where c_status not in ('0','2') and c_name ='"+name+"' order by length(c_orgcode) desc"
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        if len(data) > 0:
            return data[0][0]
        else:
            return ''

    # 保存数据到excel(不包含表头)
    def write_to_excel(self,path,date,data,templetId):
        print('保存产品"'+self.getTempleName(templetId)+'"详细信息到excel开始！')
        self.mkdir_ifnotexists(path+os.path.sep+date)
        os.chdir(path+os.path.sep+date)

        file_name = '{0}_{1}.xlsx'.format(date,self.getTempleName(templetId))
        wb = Workbook()
        sheet = wb.active
        sheet.title=date

        for r in range(len(data)):
            for c in range(len(data[r])):
                sheet.cell(row = r+1,column = c+1).value=data[r][c]
        wb.save(file_name)
        print('保存产品"'+self.getTempleName(templetId)+'"详细信息到excel结束！')

    # 保存数据到excel同时保存表头
    def write_to_excel_title(self,path,date,data,titles,templetId):
        print('保存产品详细信息到excel开始！')
        if len(data) > 0:
            self.mkdir_ifnotexists(path+os.path.sep+date)
            os.chdir(path+os.path.sep+date)
            file_name = '{0}_{1}.xlsx'.format(date,templetId)
            wb = Workbook()
            sheet = wb.active
            sheet.title=date

            for r in range(len(titles)):
                sheet.cell(row=1, column=r + 1).value = titles[r][0]

            for r in range(len(data)):
                for c in range(len(data[r])):
                    sheet.cell(row = r+2,column = c+1).value=data[r][c]
            wb.save(file_name)
        else:
            print('当天无‘'+templetId+'’产品数据！')
        print('保存产品详细信息到excel结束！')

    # 获取所有附件信息
    def read_tblobstorage_table(self,fileremotepath,filelocalpath,fundcode,startDate,tafundcode):
        print('【获取产品'+fundcode+'_'+tafundcode+'附件信息开始】')
        named_params = { 'subkeyid':fundcode}
        # sql = "SELECT  a.c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename FROM tblobstorage bs,tfundinfo a  WHERE bs.c_subkeyid=a.c_fundcode and c_catalog=:ccatalog AND c_subkeyid=:subkeyid"
        sql = " SELECT  a.c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename " \
              " FROM tblobstorage bs,tfundinfo a  " \
              " WHERE bs.c_subkeyid=a.c_fundcode  and bs.c_catalog='TFUNDINFOT' " \
              " AND c_subkeyid=:subkeyid "
        dataresult = self.cursor.execute(sql, named_params)
        data = dataresult.fetchall()
        for r in range(len(data)):
            print("\ttafundcode=%s, prdName=%s ,remotePath=%s ,localFileName=%s"%(tafundcode,data[r][0],data[r][1],data[r][2]) )
            self.getFileFromRemoteServer(filelocalpath,fileremotepath,tafundcode,data[r][0],data[r][1],data[r][2],startDate)
        print('【产品'+fundcode+'_'+tafundcode+'共获取'+str(len(data))+'条附件记录】')

    # 从远程文件服务器获取附件信息
    def getFileFromRemoteServer(self,localpath,remotepath,tafundcode,prdname,refilename,lcfilename,dataDate):
        tmpprdpath = localpath +os.path.sep + dataDate + os.path.sep + 'annex'+ os.path.sep + tafundcode
        self.mkdir_ifnotexists(tmpprdpath)
        localtmpFile = tmpprdpath+os.path.sep+lcfilename
        remotefile=remotepath+'/'+refilename
        print('\t\tremotefile='+remotefile+'\t,localtmpFile='+localtmpFile)
        try:
            self.fileSrcSftp.get(remotefile, localtmpFile)
        # shutil.copy(localfile,localtmpFile)
        except Exception as e:
            print('\t\t【获取附件异常,】' + str(e))

    # 获取存续期兑付附件信息
    def read_tblobstorage_table_Cash(self,fileremotepath,filelocalpath,lid,startDate,tafundcode):
        print('【获取产品'+tafundcode+'附件信息开始】')
        named_params = { 'csubkeyid':lid}
        # sql = "SELECT  a.c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename FROM tblobstorage bs,tfundinfo a  WHERE bs.c_subkeyid=a.c_fundcode and c_catalog=:ccatalog AND c_subkeyid=:subkeyid"
        sql = " SELECT  'nan' c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename " \
              " FROM tblobstorage bs   " \
              " WHERE (bs.c_catalog='TPDTDURATIONPROGRAM' )  " \
              " and c_subkeyid=:csubkeyid "
        print('sql2='+sql)
        dataresult = self.cursor.execute(sql, named_params)
        data = dataresult.fetchall()
        for r in range(len(data)):
            print("\ttafundcode=%s, prdName=%s ,remotePath=%s ,localFileName=%s"%(tafundcode,data[r][0],data[r][1],data[r][2]) )
            self.getFileFromRemoteServer_Cash(filelocalpath,fileremotepath,tafundcode,data[r][0],data[r][1],data[r][2],startDate)
        print('【产品'+tafundcode+'共获取'+str(len(data))+'条附件记录】')

    # 从远程文件服务器获取附件信息
    def getFileFromRemoteServer_Cash(self,localpath,remotepath,tafundcode,prdname,refilename,lcfilename,dataDate):
        tmpprdpath = localpath +os.path.sep + dataDate + os.path.sep + 'annex-cash'+ os.path.sep + tafundcode
        self.mkdir_ifnotexists(tmpprdpath)
        localtmpFile = tmpprdpath+os.path.sep+lcfilename
        remotefile=remotepath+'/'+refilename
        print('\t\tremotefile='+remotefile+'\t,localtmpFile='+localtmpFile)
        try:
            self.fileSrcSftp.get(remotefile, localtmpFile)
        # shutil.copy(localfile,localtmpFile)
        except Exception as e:
            print('\t\t【获取附件异常,】' + str(e))

    # 创建目录
    def mkdir_ifnotexists(self, path):
        path=path.strip()
        path=path.rstrip("\\")
        isExists=os.path.exists(path)

        # 判断结果
        if not isExists:
            # 如果不存在则创建目录
            os.makedirs(path)

            print('\t'+path+' 创建成功')
            return True
        else:
            # 如果目录存在则不创建，并提示目录已存在
            # print(path+' 目录已存在')
            return False

    # 从产品中心获取存续期产品
    def getDurationFundinfo(self,startDate,endDate,filelocalpath):
        print('从产品中心获取存续期startDate='+startDate+'，endDate='+endDate+'数据开始')
        sql="select  t.l_id 序号,  a.c_fundcode   产品代码, a.c_tafundcode TA产品代码, a.c_fundname 产品名称, decode( a.L_PDTTEMPLETID,'6','债权投资计划','8','资产支持计划') 产品类型," \
            " decode( t.c_type,'0','收益分配','1','赎回','2','回售','3','分期还本','4','兑付','5','清算','6','分红') 业务类型,  c_date 业务发生日期 ,  f_totalamount 业务发生总额, s_severaltime 期次," \
            " decode( c_paymenttype,'1','正常到期','2','提前还款')  兑付类型,  f_totalshare 业务发生总份数, t.d_createdate 创建日期,  f_afterfacevalue 执行后份额面值,  t.c_memo 备注" \
            " FROM tpdtdurationprogram t, tfundinfo a, member.operators o" \
            " WHERE t.c_fundcode=a.c_fundcode and  (a.L_PDTTEMPLETID in('6','8'))" \
            "   and t.c_creator = o.oper_id(+)" \
            "   and t.c_type='4'" \
            " and to_char(t.l_id) in (select biz_flow_uuid_ from jbpm4_ext_htask_ j where j.PROC_NAME_ like  '%存续期方案流程%'"\
            " and j.TASKOUTCOMEACTION = '复审通过' and j.activity_name_ = '存续期方案复审' "\
            " and to_char(j.END_, 'yyyyMMdd') >= '"+startDate+"' and to_char(j.END_, 'yyyyMMdd') <= '"+endDate+"' )"\
            " ORDER BY  t.c_type asc, t.d_createdate desc"
        dataResult = self.cursor.execute(sql)
        # 将产品信息存放到excel里
        dataRecords = dataResult.fetchall()
        self.write_to_excel_title(filelocalpath,endDate,dataRecords,dataResult.description,'存续期-兑付')
        print('从产品中心获取存续期startDate=' + startDate + '，endDate=' + endDate + '数据结束')

    # 从产品中心获取当日存续期兑付产品编码
    def getDurationFundinfo_Cash(self,templetId,startDate,endDate,filelocalpath):
        print('从产品中心获取存续期startDate='+startDate+'，endDate='+endDate+'数据开始')
        sql="select  to_char(t.l_id) as l_id,a.c_tafundcode "\
            " FROM tpdtdurationprogram t, tfundinfo a, member.operators o" \
            " WHERE t.c_fundcode=a.c_fundcode and a.L_PDTTEMPLETID= '"+templetId+"' " \
            "   and t.c_creator = o.oper_id(+)" \
            "   and t.c_type='4' " \
            " and to_char(t.l_id) in (select biz_flow_uuid_ from jbpm4_ext_htask_ j where j.PROC_NAME_ like  '%存续期方案流程%'"\
            " and j.TASKOUTCOMEACTION = '复审通过' and j.activity_name_ = '存续期方案复审' "\
            " and to_char(j.END_, 'yyyyMMdd') >= '"+startDate+"' and to_char(j.END_, 'yyyyMMdd') <= '"+endDate+"' )"\
            " ORDER BY  t.c_type asc, t.d_createdate desc"
        fundcodes=set()
        print('sql=' + sql)
        try:
            dataResult = self.cursor.execute(sql)
            # 将产品信息存放到excel里
            dataRecords = dataResult.fetchall()
            for r in range(len(dataRecords)):
                lid = dataRecords[r][0]
                ctafundcode = dataRecords[r][1]
                fundcodes.add(lid+','+ctafundcode)
            return fundcodes
            print('从产品中心获取存续期兑付startDate=' + startDate + '，endDate=' + endDate + '数据结束')
        except Exception as e:
            print('获取templetId=' + templetId + '类型的产品信息兑付异常'+str(e))
        return fundcodes

    # 从ta获取存续期产品信息
    def getFundInfoByTA(self,filelocalpath,startDate,endDate):
        print('从TA获取存续期startDate=' + startDate + '，endDate=' + endDate + '数据开始')
        sql = "SELECT a.c_fundcode 产品代码," \
                      " b.c_bondfullname 产品名称," \
                      " decode(b.c_producttype, '1', '债权', '2', '股权', '3', '资产支持计划') 产品类型," \
                      " decode(c_rdmflag, 'l', '回售', 'm', '赎回', 'i', '分期还本') 业务类型," \
                      " a.d_rdmdate 业务发生日期," \
                      " a.d_impdate 导入日期," \
                      " b.f_issueprice 执行后份额面值," \
                      " a.sumbalance 业务发生总额," \
                      " a.sumshare 业务发生总份数" \
            " FROM (SELECT a.c_fundcode, a.c_rdmflag, a.d_rdmdate, a.d_impdate," \
                      "    sum(a.f_redeembalance) sumbalance, sum(a.f_redeemshares) sumshare" \
                    " FROM tbatchrightline a" \
                    " group by a.c_fundcode, a.c_rdmflag, a.d_rdmdate, a.d_impdate) a, tbondinfo b" \
            " where a.c_fundcode = b.c_bondcode " \
              " and a.d_impdate >= '"+startDate+"' and a.d_impdate <= '"+endDate+"'" \
            " union all" \
            " SELECT a.c_fundcode 产品代码," \
                    " b.c_bondfullname 产品名称," \
                    " decode(b.c_producttype, '1', '债权', '2', '股权', '3', '资产支持计划') 产品类型," \
                    " '收益分配' 业务类型," \
                    " a.d_regdate 业务发生日期," \
                    " a.d_cdate 导入日期," \
                    " b.f_issueprice 执行后份额面值," \
                    " a.sumbalance 业务发生总额," \
                    " a.sumshare 业务发生总份数" \
            " FROM (SELECT a.c_fundcode, a.d_regdate, a.d_cdate," \
                    " sum(a.f_realbalance) sumbalance, sum(a.f_realshares) sumshare" \
                    " FROM tdividenddetail a" \
                    " group by a.c_fundcode, a.d_regdate, a.d_cdate) a,tbondinfo b" \
            " where a.c_fundcode = b.c_bondcode " \
                    " and a.d_cdate >= '"+startDate+"' and a.d_cdate <= '"+endDate+"'" \

        dataResult = self.cursor.execute(sql)
        # 将产品信息存放到excel里
        dataRecords = dataResult.fetchall()
        self.write_to_excel_title(filelocalpath, endDate, dataRecords, dataResult.description, '存续期-无兑付')
        print('从TA获取存续期startDate=' + startDate + '，endDate=' + endDate + '数据结束')
    ########################################################################################################################
    def rta_disconnect(self):
        try:
            if self.dbrta is not None:
                self.dbrta.close()
                print("Rta Connection terminated.")
        except cx_Oracle.DatabaseError as e:
            print("close db  except." + str(e))

    def getDbRtaCursor(self,dbIp,dbPort,dbSID,dbUserName,dbPassword):
        try:
            dsn_tns = cx_Oracle.makedsn(dbIp, dbPort, service_name=dbSID)
            self.dbrta = cx_Oracle.connect(dbUserName, dbPassword, dsn_tns)
            cursor = self.dbrta.cursor()
            savelog("connet %s:%s/%s database(rta) successful!" % (dbIp,dbPort,dbSID))
            return cursor
        except cx_Oracle.Error as e:
            savelog("init connet database(rta) failed,exit!" + str(e))
            time.sleep(5)
            sys.exit(-1)

    def set_rtacursor(self,dbip,dbport,dbSID,dbUserName,dbPassword):
        #读取rta
        self.rtacursor = self.getDbRtaCursor(dbip,dbport,dbSID,dbUserName,dbPassword)

    def get_total_reg_scope(self,start_date,end_date):
        #获得总体情况登记规模
        sql = '''
            select sum(c.F_PRDREGBALA)
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m
                WHERE a.C_PDTTYPE=1 AND a.L_PDTTEMPLETID = 8 AND a.C_AUDITFLAG in(3,4,5,6,7,12,13,15,16,20,21,22,25,29,55,56,57,58) AND  a.c_fundcode = b.c_fundcode
                AND a.c_fundcode = c.c_fundcode
                AND c.c_mgrorgno = d.mem_code(+)
                AND a.c_fundcode = e.c_fundcode(+)
                AND a.c_fundcode = h.c_fundcode(+)
                AND h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')<='enddate'
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("总体情况登记规模,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:" + str(data),False)
        total = 0.0
        for item in data:
            if item[0] is not None and item[0] != "None":
                total = total + float(item[0])

        return str(total)
    
    def get_total_reg_count(self,start_date,end_date):
        #获得总体情况登记只数
        sql = '''
            select count(a.c_fundcode)
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m
                WHERE a.C_PDTTYPE=1 AND a.L_PDTTEMPLETID = 8 AND a.C_AUDITFLAG in(3,4,5,6,7,12,13,15,16,20,21,22,25,29,55,56,57,58) AND  a.c_fundcode = b.c_fundcode
                AND a.c_fundcode = c.c_fundcode
                AND c.c_mgrorgno = d.mem_code(+)
                AND a.c_fundcode = e.c_fundcode(+)
                AND a.c_fundcode = h.c_fundcode(+)
                AND h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')<='enddate'
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("总体情况登记只数,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(data),False)
        total = 0
        for item in data:
            total = total + int(item[0])

        return str(total)

    def get_total_release_scope(self,start_date,end_date):
        #获得总体情况的x年x月发行规模
        sql = '''
            select a.c_fundcode,c.c_prdisclass,c.C_PRDISORDER 
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m,
                    (select m.c_fundcode, m.c_stagefundcode, n.c_fundname as c_stagefundname, n.c_tafundcode as c_stagetafundcode
                        FROM fundcrm.tpdtstagerelation m, fundcrm.tfundinfo n
                        WHERE m.c_stagefundcode = n.c_fundcode) f
                WHERE a.L_PDTTEMPLETID = 8 AND (a.C_AUDITFLAG in(5,6,7,15,16,21,57,58)) AND  a.c_fundcode = b.c_fundcode
                and a.c_fundcode = c.c_fundcode
                and c.c_mgrorgno = d.mem_code(+)
                and a.c_fundcode = f.c_fundcode(+)
                and a.c_fundcode = e.c_fundcode(+)
                and a.c_fundcode = h.c_fundcode(+)
                and h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')<='enddate'
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("获取产品代码,分级标识,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(data),False)
        total = 0.0
        for item in data:
            #遍历每一项
            sqlsub = ""
            if (item[1] is not None and int(item[1]) == 0 and item[2] is not None and int(item[2]) == 0) or \
                (item[1] is not None and int(item[1]) == 0 and item[2] is None):
                #发行规模（c_prdisclass=0 && C_PRDISORDER=0）
                sqlsub = "SELECT F_BATCHFACTPAYBALA FROM fundcrm.tpdtbatchinfo o WHERE o.C_FUNDCODE = '" + str(item[0]) + "'"
                savelog("发行规模:" + sqlsub)
            elif item[1] is not None and int(item[1]) == 1:
                #分级产品发行规模（c_prdisclass=1）
                sqlsub = "SELECT sum(F_CLASSFACTBALE) FROM fundcrm.TPDTPRDCLASSINFO o WHERE o.C_FUNDCODE ='" + str(item[0]) + "'"
                savelog("分级产品发行规模:" + sqlsub)
            elif item[2] is not None and int(item[2]) == 1:
                #分次产品发行规模(C_PRDISORDER=1)
                sqlsub = "SELECT sum(F_ORDERFACTCOLLECTBALA) FROM fundcrm.TPDTORDERINFO o WHERE o.C_FUNDCODE = '" + str(item[0]) + "'"
                savelog("分次产品发行规模:" + sqlsub)
            else:
                savelog("总体情况,获取产品代码,该分级标识参数不进行处理" + str(item[0]) + " " + str(item[1]) + " " + str(item[2]))
                continue
            
            dataresult_sub = self.cursor.execute(sqlsub)
            data_sub = dataresult_sub.fetchall()
            savelog("查询结果:" + str(data_sub),False)

            total = total + float(data_sub[0][0])

        return str(total)

    def get_total_release_count(self,start_date,end_date):
        #获得总体情况发行只数
        sql = '''
            select count(a.c_fundcode) 
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m,
                    (select m.c_fundcode, m.c_stagefundcode, n.c_fundname as c_stagefundname, n.c_tafundcode as c_stagetafundcode
                        FROM fundcrm.tpdtstagerelation m, fundcrm.tfundinfo n
                        WHERE m.c_stagefundcode = n.c_fundcode) f
                WHERE a.L_PDTTEMPLETID = 8 AND (a.C_AUDITFLAG in(5,6,7,15,16,21,57,58)) AND  a.c_fundcode = b.c_fundcode
                and a.c_fundcode = c.c_fundcode
                and c.c_mgrorgno = d.mem_code(+)
                and a.c_fundcode = f.c_fundcode(+)
                and a.c_fundcode = e.c_fundcode(+)
                and a.c_fundcode = h.c_fundcode(+)
                and h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')<='enddate'
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("总体情况发行只数,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:" + str(data),False)
        total = 0
        for item in data:
            total = total + int(item[0])

        return str(total)

    def get_trustee_reg_scope(self,start_date,end_date):
        #获得受托人登记规模
        sql = '''
            select d.mem_name,sum(c.F_PRDREGBALA)
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m
                WHERE a.C_PDTTYPE=1 AND a.L_PDTTEMPLETID = 8 AND a.C_AUDITFLAG in(3,4,5,6,7,12,13,15,16,20,21,22,25,29,55,56,57,58) AND  a.c_fundcode = b.c_fundcode
                AND a.c_fundcode = c.c_fundcode
                AND c.c_mgrorgno = d.mem_code(+)
                AND a.c_fundcode = e.c_fundcode(+)
                AND a.c_fundcode = h.c_fundcode(+)
                AND h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')<='enddate' 
                GROUP BY  d.mem_name
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("受托人登记规模,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:" + str(data),False)
        resmap = {}
        line = 0
        for item in data:
            resmap[item[0]] = [str(item[1]),line]
            #保存mem_name对应的值和它所在的行数
            line = line + 1

        return resmap
    
    def get_trustee_reg_count(self,start_date,end_date):
        #获得受托人登记只数
        sql = '''
            select d.mem_name,count(d.mem_name)
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m
                WHERE a.C_PDTTYPE=1 AND a.L_PDTTEMPLETID = 8 AND a.C_AUDITFLAG in(3,4,5,6,7,12,13,15,16,20,21,22,25,29,55,56,57,58) AND  a.c_fundcode = b.c_fundcode
                AND a.c_fundcode = c.c_fundcode
                AND c.c_mgrorgno = d.mem_code(+)
                AND a.c_fundcode = e.c_fundcode(+)
                AND a.c_fundcode = h.c_fundcode(+)
                AND h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')<='enddate' 
                GROUP BY  d.mem_name
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("受托人登记只数,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:" + str(data),False)
        line = 0
        resmap = {}
        for item in data:
            resmap[item[0]] = [str(item[1]),line]
            #保存mem_name对应的值和它所在的行数
            line = line + 1

        return resmap

    def get_trustee_release_scope(self,start_date,end_date):
        #获得受托人发行规模
        sql = '''
            select DISTINCT c.c_mgrorgno,d.mem_name
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m,
                    (select m.c_fundcode, m.c_stagefundcode, n.c_fundname as c_stagefundname, n.c_tafundcode as c_stagetafundcode
                        FROM fundcrm.tpdtstagerelation m, fundcrm.tfundinfo n
                        WHERE m.c_stagefundcode = n.c_fundcode) f
                WHERE a.L_PDTTEMPLETID = 8 AND (a.C_AUDITFLAG in(5,6,7,15,16,21,57,58)) AND  a.c_fundcode = b.c_fundcode
                and a.c_fundcode = c.c_fundcode
                and c.c_mgrorgno = d.mem_code(+)
                and a.c_fundcode = f.c_fundcode(+)
                and a.c_fundcode = e.c_fundcode(+)
                and a.c_fundcode = h.c_fundcode(+)
                and h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')<='enddate'
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("受托人发行规模,受托人机构号,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:" + str(data),False)
        line = 0
        resmap = {}
        for item in data:
            c_mgrorgno = item[0]
            mem_name = item[1]
            sql_mgrorgno = '''
                select a.c_fundcode,c.c_prdisclass,c.C_PRDISORDER 
                    FROM fundcrm.tfundinfo a,
                        fundcrm.tpdtoperateinfo b,
                        fundcrm.tpdtbasicinfo c,
                        member.member_info d,
                        fundcrm.tpdtextendinfo e,
                        fundcrm.tpdtbatchinfo h,
                        member.member_info m,
                        (select m.c_fundcode, m.c_stagefundcode, n.c_fundname as c_stagefundname, n.c_tafundcode as c_stagetafundcode
                            FROM fundcrm.tpdtstagerelation m, fundcrm.tfundinfo n
                            WHERE m.c_stagefundcode = n.c_fundcode) f
                    WHERE a.L_PDTTEMPLETID = 8 AND (a.C_AUDITFLAG in(5,6,7,15,16,21,57,58)) AND  a.c_fundcode = b.c_fundcode
                    and a.c_fundcode = c.c_fundcode
                    and c.c_mgrorgno = d.mem_code(+)
                    and a.c_fundcode = f.c_fundcode(+)
                    and a.c_fundcode = e.c_fundcode(+)
                    and a.c_fundcode = h.c_fundcode(+)
                    and h.c_trustee = m.mem_code(+)
                    AND c.c_mgrorgno = trustee_org_code
                    AND TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')<='enddate'
            '''
            sql_mgrorgno = sql_mgrorgno.replace("startdate",start_date).replace("enddate",end_date).replace("trustee_org_code",c_mgrorgno)
            savelog("获取产品代码、分级标识(0代表非分级,1代表分级),分次标识(0代表非分次,1代表分次) sql:" + sql_mgrorgno,False)
            data_mgrorno_result = self.cursor.execute(sql_mgrorgno)
            data_mgrorno = data_mgrorno_result.fetchall()
            savelog("查询结果:" + str(data_mgrorno))
            total = 0.0
            for item_mgrorgno in data_mgrorno:
                sqlsub = ""
                if (item_mgrorgno[1] is not None and int(item_mgrorgno[1]) == 0 and item_mgrorgno[2] is not None and int(item_mgrorgno[2]) == 0) or \
                    (item_mgrorgno[1] is not None and int(item_mgrorgno[1]) == 0 and item_mgrorgno[2] is None):
                    #发行规模（c_prdisclass=0 && C_PRDISORDER=0）
                    sqlsub = "SELECT F_BATCHFACTPAYBALA FROM fundcrm.tpdtbatchinfo o WHERE o.C_FUNDCODE = '" + str(item_mgrorgno[0]) + "'"
                    savelog("发行规模:" + sqlsub)
                elif item_mgrorgno[1] is not None and int(item_mgrorgno[1]) == 1:
                    #分级产品发行规模（c_prdisclass=1）
                    sqlsub = "SELECT sum(F_CLASSFACTBALE) FROM fundcrm.TPDTPRDCLASSINFO o WHERE o.C_FUNDCODE ='" + str(item_mgrorgno[0]) + "'"
                    savelog("分级产品发行规模:" + sqlsub)
                elif item_mgrorgno[2] is not None and int(item_mgrorgno[2]) == 1:
                    #分次产品发行规模(C_PRDISORDER=1)
                    sqlsub = "SELECT sum(F_ORDERFACTCOLLECTBALA) FROM fundcrm.TPDTORDERINFO o WHERE o.C_FUNDCODE = '" + str(item_mgrorgno[0]) + "'"
                    savelog("分次产品发行规模:" + sqlsub)
                else:
                    savelog("受托人,获取产品代码,该分级标识参数不进行处理 " + str(item_mgrorgno[0]) + " " + str(item_mgrorgno[1]) + " " + str(item_mgrorgno[2]))
                    continue

                dataresult_sub = self.cursor.execute(sqlsub)
                data_sub = dataresult_sub.fetchall()
                savelog("查询结果:" + str(data_sub),False)

                total = total + float(data_sub[0][0])

            resmap[mem_name] = [str(total),line]
            line = line + 1

        return resmap

    def get_trustee_release_count(self,start_date,end_date):
        #获得受托人当前年月发行只数
        sql = '''
            select d.mem_name,count(d.mem_name)
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m,
                    (select m.c_fundcode, m.c_stagefundcode, n.c_fundname as c_stagefundname, n.c_tafundcode as c_stagetafundcode
                        FROM fundcrm.tpdtstagerelation m, fundcrm.tfundinfo n
                        WHERE m.c_stagefundcode = n.c_fundcode) f
                WHERE a.L_PDTTEMPLETID = 8 AND (a.C_AUDITFLAG in(5,6,7,15,16,21,57,58)) AND  a.c_fundcode = b.c_fundcode
                and a.c_fundcode = c.c_fundcode
                and c.c_mgrorgno = d.mem_code(+)
                and a.c_fundcode = f.c_fundcode(+)
                and a.c_fundcode = e.c_fundcode(+)
                and a.c_fundcode = h.c_fundcode(+)
                and h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')<='enddate'
                GROUP BY  d.mem_name
            '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("受托人发行只数,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        
        line = 0
        resmap = {}
        for item in data:
            resmap[item[0]] = [str(item[1]),line]
            line = line + 1

        return resmap

    def get_deposit_bank_reg_scope(self,start_date,end_date):
        #获得托管银行登记规模
        sql = '''
            select m.mem_name,sum(c.F_PRDREGBALA)
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m
                WHERE a.C_PDTTYPE=1 AND a.L_PDTTEMPLETID = 8 AND a.C_AUDITFLAG in(3,4,5,6,7,12,13,15,16,20,21,22,25,29,55,56,57,58) AND  a.c_fundcode = b.c_fundcode
                and a.c_fundcode = c.c_fundcode
                and c.c_mgrorgno = d.mem_code(+)
                and a.c_fundcode = e.c_fundcode(+)
                and a.c_fundcode = h.c_fundcode(+)
                and h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')<='enddate' 
                GROUP BY  m.mem_name
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("托管银行登记规模,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(data),False)
        line = 0
        resmap = {}
        for item in data:
            resmap[item[0]] = [str(item[1]),line]
            line = line + 1

        return resmap
    
    def get_depoist_bank_reg_count(self,start_date,end_date):
        #获得托管银行登记只数
        sql = '''
            select m.mem_name,count(m.mem_name)
            FROM fundcrm.tfundinfo a,
                fundcrm.tpdtoperateinfo b,
                fundcrm.tpdtbasicinfo c,
                member.member_info d,
                fundcrm.tpdtextendinfo e,
                fundcrm.tpdtbatchinfo h,
                member.member_info m
            WHERE a.C_PDTTYPE=1 AND a.L_PDTTEMPLETID = 8 AND a.C_AUDITFLAG in(3,4,5,6,7,12,13,15,16,20,21,22,25,29,55,56,57,58) AND  a.c_fundcode = b.c_fundcode
            and a.c_fundcode = c.c_fundcode
            and c.c_mgrorgno = d.mem_code(+)
            and a.c_fundcode = e.c_fundcode(+)
            and a.c_fundcode = h.c_fundcode(+)
            and h.c_trustee = m.mem_code(+)
            AND TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_preregisterdate,'yyyy-mm-dd')<='enddate' 
            GROUP BY  m.mem_name
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("托管银行登记只数,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(data),False)
        line = 0
        resmap = {}
        for item in data:
            resmap[item[0]] = [str(item[1]),line]
            line = line + 1

        return resmap
    
    def get_deposit_bank_release_scope(self,start_date,end_date):
        #获得托管银行发行规模
        #查询托管银行机构号
        sql = '''
            select DISTINCT h.c_trustee,m.mem_name
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m,
                    (select m.c_fundcode, m.c_stagefundcode, n.c_fundname as c_stagefundname, n.c_tafundcode as c_stagetafundcode
                        FROM fundcrm.tpdtstagerelation m, fundcrm.tfundinfo n
                        WHERE m.c_stagefundcode = n.c_fundcode) f
                WHERE a.L_PDTTEMPLETID = 8 AND (a.C_AUDITFLAG in(5,6,7,15,16,21,57,58)) AND  a.c_fundcode = b.c_fundcode
                and a.c_fundcode = c.c_fundcode
                and c.c_mgrorgno = d.mem_code(+)
                and a.c_fundcode = f.c_fundcode(+)
                and a.c_fundcode = e.c_fundcode(+)
                and a.c_fundcode = h.c_fundcode(+)
                and h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')<='enddate'
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("托管银行发行规模,查询托管银行机构号,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("托管银行机构号查询结果:" + str(data))
        
        line = 0
        resmap = {}
        for item in data:
            c_trustee = item[0]
            mem_name = item[1]
            sql_c_trustee = '''
                            select a.c_fundcode,c.c_prdisclass,c.C_PRDISORDER 
                                FROM fundcrm.tfundinfo a,
                                    fundcrm.tpdtoperateinfo b,
                                    fundcrm.tpdtbasicinfo c,
                                    member.member_info d,
                                    fundcrm.tpdtextendinfo e,
                                    fundcrm.tpdtbatchinfo h,
                                    member.member_info m,
                                    (select m.c_fundcode, m.c_stagefundcode, n.c_fundname as c_stagefundname, n.c_tafundcode as c_stagetafundcode
                                        FROM fundcrm.tpdtstagerelation m, fundcrm.tfundinfo n
                                        WHERE m.c_stagefundcode = n.c_fundcode) f
                                WHERE a.L_PDTTEMPLETID = 8 AND (a.C_AUDITFLAG in(5,6,7,15,16,21,57,58)) AND  a.c_fundcode = b.c_fundcode
                                and a.c_fundcode = c.c_fundcode
                                and c.c_mgrorgno = d.mem_code(+)
                                and a.c_fundcode = f.c_fundcode(+)
                                and a.c_fundcode = e.c_fundcode(+)
                                and a.c_fundcode = h.c_fundcode(+)
                                and h.c_trustee = m.mem_code(+)
                                AND h.c_trustee = 'c_trustee_code'
                                AND TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')<='enddate'
                            '''
            sql_c_trustee = sql_c_trustee.replace("startdate",start_date).replace("enddate",end_date).replace("c_trustee_code",c_trustee)
            savelog("获取产品代码,分级标识:" + sql_c_trustee)
            dataresult_trustee = self.cursor.execute(sql_c_trustee)
            data_trustee = dataresult_trustee.fetchall()
            savelog("产品代码,分级标识查询结果:" + str(data_trustee))
            total = 0.0
            for item_trustee in data_trustee:
                sqlsub = ""
                if (item_trustee[1] is not None and int(item_trustee[1]) == 0 and item_trustee[2] is not None and int(item_trustee[2]) == 0) or \
                    (item_trustee[1] is not None and int(item_trustee[1]) == 0 and item_trustee[2] is None):
                    #发行规模（c_prdisclass=0 && C_PRDISORDER=0）
                    sqlsub = "SELECT F_BATCHFACTPAYBALA FROM fundcrm.tpdtbatchinfo o WHERE o.C_FUNDCODE = '" + str(item_trustee[0]) + "'"
                    savelog("发行规模:" + sqlsub)
                elif item_trustee[1] is not None and int(item_trustee[1]) == 1:
                    #分级产品发行规模（c_prdisclass=1）
                    sqlsub = "SELECT sum(F_CLASSFACTBALE) FROM fundcrm.TPDTPRDCLASSINFO o WHERE o.C_FUNDCODE ='" + str(item_trustee[0]) + "'"
                    savelog("分级产品发行规模:" + sqlsub)
                elif item_trustee[2] is not None and int(item_trustee[2]) == 1:
                    #分次产品发行规模(C_PRDISORDER=1)
                    sqlsub = "SELECT sum(F_ORDERFACTCOLLECTBALA) FROM fundcrm.TPDTORDERINFO o WHERE o.C_FUNDCODE = '" + str(item_trustee[0]) + "'"
                    savelog("分次产品发行规模:" + sqlsub)
                else:
                    savelog("托管银行,获取产品代码,该分级标识参数不进行处理 " + str(item_trustee[0]) + " " + str(item_trustee[1]) + " " + str(item_trustee[2]))
                    continue

                dataresult_sub = self.cursor.execute(sqlsub)
                data_sub = dataresult_sub.fetchall()
                savelog("查询结果:" + str(data_sub),False)

                total = total + float(data_sub[0][0])
            
            resmap[mem_name] = [str(total),line]
            line = line + 1

        return resmap

    def get_deposit_bank_release_count(self,start_date,end_date):
        #获得托管银行年月发行只数
        sql = '''
            select m.mem_name,count(m.mem_name)
                FROM fundcrm.tfundinfo a,
                    fundcrm.tpdtoperateinfo b,
                    fundcrm.tpdtbasicinfo c,
                    member.member_info d,
                    fundcrm.tpdtextendinfo e,
                    fundcrm.tpdtbatchinfo h,
                    member.member_info m,
                    (select m.c_fundcode, m.c_stagefundcode, n.c_fundname as c_stagefundname, n.c_tafundcode as c_stagetafundcode
                        FROM fundcrm.tpdtstagerelation m, fundcrm.tfundinfo n
                        WHERE m.c_stagefundcode = n.c_fundcode) f
                WHERE a.L_PDTTEMPLETID = 8 AND (a.C_AUDITFLAG in(5,6,7,15,16,21,57,58)) AND  a.c_fundcode = b.c_fundcode
                and a.c_fundcode = c.c_fundcode
                and c.c_mgrorgno = d.mem_code(+)
                and a.c_fundcode = f.c_fundcode(+)
                and a.c_fundcode = e.c_fundcode(+)
                and a.c_fundcode = h.c_fundcode(+)
                and h.c_trustee = m.mem_code(+)
                AND TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(b.d_accessstoredate,'yyyy-mm-dd')<='enddate'
                GROUP BY  m.mem_name
        '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("托管银行发行只数,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(data),False)
        line = 0
        resmap = {}
        for item in data:
            resmap[item[0]] = [str(item[1]),line]
            line = line + 1

        return resmap

    def get_amount_total_scope(self,end_date):
        #获得历史累计总体情况的存量规模
        ta4sql = '''
            SELECT
                sum(a.f_lastshares * b.price)
                    FROM
                        (
                        SELECT
                            a.c_fundacco,
                            a.c_fundcode,
                            a.d_cdate d_cdate,
                            a.f_occurshares,
                            a.f_lastshares,
                            a.d_requestdate d_date,
                            ROW_NUMBER() OVER (PARTITION BY a.c_fundacco,a.c_fundcode ORDER BY a.d_requestdate desc,a.L_SERIALNO desc) AS rn
                        FROM
                            tsharecurrents a
                        WHERE
                            a.C_OUTBUSINFLAG  NOT IN ('031','032') and
                            a.d_cdate <= enddate ) a ,
                        (
                        SELECT
                                c_bondcode fundcode,
                                C_BONDFULLNAME fundname,
                                c_mgrcode,
                                c_producttype producttype,
                                f_issueprice price
                        FROM
                                tbondinfo a
                        WHERE
                    a.c_producttype = '3' ) b
                    WHERE
                        a.rn = 1
                        AND a.c_fundcode = b.fundcode
        '''
        sql = '''
            SELECT
                sum(a.f_lastshares * b.price)
            FROM
                (
                SELECT
                    a.c_fundacco,
                    a.c_fundcode,
                    to_char(a.d_cdate, 'yyyymmdd') d_cdate,
                    a.f_occurshares,
                    a.f_lastshares,
                    to_char(a.d_requestdate, 'yyyymmdd') d_date,
                    ROW_NUMBER() OVER (PARTITION BY a.c_fundacco,a.c_fundcode ORDER BY to_char(a.d_requestdate, 'yyyymmdd') desc,a.L_SERIALNO desc) AS rn
                FROM
                    tsharecurrents a
                WHERE
                    a.C_OUTBUSINFLAG  NOT IN ('031','032') and
                    to_char(a.d_cdate, 'yyyymmdd') <= 'enddate' ) a ,
                (
                SELECT
                        c_bondcode fundcode,
                        C_BONDFULLNAME fundname,
                        c_mgrcode,
                        c_producttype producttype,
                        f_issueprice price
                FROM
                        tbondinfo a
                WHERE
            a.c_producttype = '3' ) b
            WHERE
                a.rn = 1
                AND a.c_fundcode = b.fundcode
        '''
        end_date = end_date.replace("-","")
        # sql = sql.replace("enddate",end_date)
        sql = ta4sql.replace("enddate",end_date)
        savelog("历史累计的总体情况存量规模,截至日期:" + end_date + " " + sql,False)
        dataresult = self.rtacursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(data),False)
        total = 0.0
        for item in data:
            if item[0] is not None and item[0] != "None":
                total = total + float(item[0])

        return str(total)

    def get_amount_trustee_scope(self,end_date):
        #获得历史累计受托人的存量规模
        ta4sql = '''
            SELECT
            c.c_memname,
                sum(a.f_lastshares * b.price)
                FROM
                    (
                    SELECT
                        a.c_fundacco,
                        a.c_fundcode,
                        a.d_cdate d_cdate,
                        a.f_occurshares,
                        a.f_lastshares,
                        a.d_requestdate d_date,
                        ROW_NUMBER() OVER (PARTITION BY a.c_fundacco,a.c_fundcode ORDER BY a.d_requestdate desc,a.L_SERIALNO desc) AS rn
                    FROM
                        tsharecurrents a
                    WHERE
                        a.C_OUTBUSINFLAG  NOT IN ('031','032') and
                        a.d_cdate <= enddate ) a ,
                    (
                    SELECT
                            c_bondcode fundcode,
                            C_BONDFULLNAME fundname,
                            c_mgrcode,
                            c_producttype producttype,
                            f_issueprice price
                    FROM
                            tbondinfo a
                    WHERE
                        a.c_producttype = '3' ) b,
                    (
                    SELECT
                        a.c_custno c_fundacco,
                        m.c_memname c_memname
                    FROM
                        rta.tcustomerinfo a ,
                        tmemberinfo m
                    WHERE
                a.c_custodiancode = m.c_memcode
                ) c
                WHERE
                        a.rn = 1
                    AND a.c_fundcode = b.fundcode(+)
                    AND a.c_fundacco = c.c_fundacco(+)
                    GROUP BY c.c_memname
        '''
        sql = '''
            SELECT
            c.c_memname,
                sum(a.f_lastshares * b.price)
            FROM
                (
                SELECT
                    a.c_fundacco,
                    a.c_fundcode,
                    to_char(a.d_cdate, 'yyyymmdd') d_cdate,
                    a.f_occurshares,
                    a.f_lastshares,
                    to_char(a.d_requestdate, 'yyyymmdd') d_date,
                    ROW_NUMBER() OVER (PARTITION BY a.c_fundacco,a.c_fundcode ORDER BY to_char(a.d_requestdate, 'yyyymmdd') desc,a.L_SERIALNO desc) AS rn
                FROM
                    tsharecurrents a
                WHERE
                    a.C_OUTBUSINFLAG  NOT IN ('031','032') and
                    to_char(a.d_cdate, 'yyyymmdd') <= 'enddate' ) a ,
                (
                SELECT
                        c_bondcode fundcode,
                        C_BONDFULLNAME fundname,
                        c_mgrcode,
                        c_producttype producttype,
                        f_issueprice price
                FROM
                        tbondinfo a
                WHERE
                    a.c_producttype = '3' ) b,
                (
                SELECT
                    a.c_custno c_fundacco,
                    m.c_memname c_memname
                FROM
                    tcustomerinfo a ,
                    tmemberinfo m
                WHERE
            a.c_custodiancode = m.c_memcode
            ) c
            WHERE
                    a.rn = 1
                AND a.c_fundcode = b.fundcode(+)
	            AND a.c_fundacco = c.c_fundacco(+)
                GROUP BY c.c_memname
            '''
        end_date = end_date.replace("-","")
        # sql = sql.replace("enddate",end_date)
        sql = ta4sql.replace("enddate",end_date)
        savelog("历史累计的受托人存量规模,截至日期:" + end_date + " " + sql,False)
        dataresult = self.rtacursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:" + str(len(data)) + "条数据",False)
        line = 0
        resmap = {}
        total = 0.0
        for item in data:
            k = ""
            v = 0.0
            if item[0] is not None:
                k = item[0]
            else:
                k = "未知" + str(line)

            if item[1] is not None:
                v = float(item[1])

            savelog(k + " " + str(v))
            resmap[k] = [str(v),line]
            total = total + v
            line = line + 1

        savelog("历史累计的受托人存量规模，数据条数:" + str(len(data)) + " amount:" + str(total))
        return resmap

    def get_amount_deposit_bank_scope(self,end_date):
        #获得历史累计托管银行的存量规模
        ta4sql = '''
            SELECT
                b.c_memname,
                    sum(a.f_lastshares * b.price)
                FROM
                    (
                    SELECT
                        a.c_fundacco,
                        a.c_fundcode,
                        a.d_cdate d_cdate,
                        a.f_occurshares,
                        a.f_lastshares,
                        a.d_requestdate d_date,
                        ROW_NUMBER() OVER (PARTITION BY a.c_fundacco,a.c_fundcode ORDER BY a.d_requestdate desc,a.L_SERIALNO desc) AS rn
                    FROM
                        tsharecurrents a
                    WHERE
                        a.C_OUTBUSINFLAG  NOT IN ('031','032') and
                        a.d_cdate <= enddate ) a ,
                    (
                    SELECT
                            c_bondcode fundcode,
                            C_BONDFULLNAME fundname,
                            c_mgrcode,
                            c_producttype producttype,
                            f_issueprice price,
                        m.c_memname c_memname
                    FROM
                            tbondinfo a,tmemberinfo m
                    WHERE
                        a.c_mgrcode = m.c_memcode
                        AND a.c_producttype = '3' ) b
                WHERE
                        a.rn = 1
                    AND a.c_fundcode = b.fundcode
                    GROUP BY b.c_memname
        '''
        sql = '''
            SELECT
                b.c_memname,
                    sum(a.f_lastshares * b.price)
                FROM
                    (
                    SELECT
                        a.c_fundacco,
                        a.c_fundcode,
                        to_char(a.d_cdate, 'yyyymmdd') d_cdate,
                        a.f_occurshares,
                        a.f_lastshares,
                        to_char(a.d_requestdate, 'yyyymmdd') d_date,
                        ROW_NUMBER() OVER (PARTITION BY a.c_fundacco,a.c_fundcode ORDER BY to_char(a.d_requestdate, 'yyyymmdd') desc,a.L_SERIALNO desc) AS rn
                    FROM
                        tsharecurrents a
                    WHERE
                        a.C_OUTBUSINFLAG  NOT IN ('031','032') and
                        to_char(a.d_cdate, 'yyyymmdd') <= 'enddate' ) a ,
                    (
                    SELECT
                            c_bondcode fundcode,
                            C_BONDFULLNAME fundname,
                            c_mgrcode,
                            c_producttype producttype,
                            f_issueprice price,
                        m.c_memname c_memname
                    FROM
                            tbondinfo a,tmemberinfo m
                    WHERE
                        a.c_mgrcode = m.c_memcode
                        AND a.c_producttype = '3' ) b
                WHERE
                        a.rn = 1
                    AND a.c_fundcode = b.fundcode
                    GROUP BY b.c_memname
            '''
        end_date = end_date.replace("-","")
        # sql = sql.replace("enddate",end_date)
        sql = ta4sql.replace("enddate",end_date)
        savelog("历史累计的托管银行存量规模,截至日期:" + end_date + " " + sql,False)
        dataresult = self.rtacursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:" + str(len(data)) + "条数据",False)
        line = 0
        resmap = {}
        total = 0.0
        for item in data:
            k = ""
            v = 0.0
            if item[0] is not None:
                k = item[0]
            if item[1] is not None:
                v = float(item[1])

            savelog(k + " " + str(v))
            resmap[k] = [str(v),line]
            line = line + 1
            total = total + v

        savelog("历史累计的托管银行存量规模，数据条数:" + str(len(data)) + " amount:" + str(total))
        return resmap

    def create_asset_plan_reg_file(self,filepath,y="",m="",d=""):
        #产生中保登资产支持计划登记情况表.xlsx
        savelog("create_asset_plan_reg_file filepath:" + filepath)
        datemap = get_date_info_map(y,m,d)
        year = datemap["year"]
        month = datemap["month"]
        day = datemap["day"]
        quarter = datemap["quarter"]
        month_date_start = datemap["month_date_start"]
        month_date_end = datemap["month_date_end"]
        pre_year_month_start = datemap["pre_year_month_start"]
        pre_year_month_end = datemap["pre_year_month_end"]

        quarter_date_start = datemap["quarter_date_start"]
        quarter_date_end = datemap["quarter_date_end"]
        pre_year_quarter_date_start = datemap["pre_year_quarter_date_start"]
        pre_year_quarter_date_end = datemap["pre_year_quarter_date_end"]

        year_date_start = datemap["year_date_start"]
        year_date_end = datemap["year_date_end"]
        pre_year_date_start = datemap["pre_year_date_start"]
        pre_year_date_end = datemap["pre_year_date_end"]

        amount_date_init = datemap["init_date"]
        amount_date_cur = datemap["cur_date"]
        pre_year_cur_date = datemap["pre_year_cur_date"]

        total = "总体情况"
        trustee = "受托人"
        deposit_bank = "托管银行"
        t1 = year + "年" + month + "月登记规模"
        t2 = "去年" + month + "月登记规模"
        t3 = year + "年" + month + "月登记只数"
        t4 = "去年" + month + "月登记只数"
        t5 = year + "年" + month + "月发行规模"
        t6 = "去年" + month + "月发行规模"
        t7 = year + "年" + month + "月发行只数"
        t8 = "去年" + month + "月发行只数"
        t9 = year + "年" + quarter + "季度登记规模"
        t10 = "去年" + quarter + "季度登记规模"
        t11 = year + "年" + quarter + "季度登记只数"
        t12 = "去年" + quarter + "季度登记只数"
        t13 = year + "年" + quarter + "季度发行规模"
        t14 = "去年" + quarter + "季度发行规模"
        t15 = year + "年" + quarter + "季度发行只数"
        t16 = "去年" + quarter + "季度发行只数"
        t17 = year + "年累计登记规模"
        t18 = "去年累计登记规模"
        t19 = year + "年累计登记只数"
        t20 = "去年累计登记只数"
        t21 = year + "年累计发行规模"
        t22 = "去年累计发行规模"
        t23 = year + "年累计发行只数"
        t24 = "去年累计发行只数"
        t25 = "历史累计登记规模（截至" + year +"年" + month + "月" + day + "日）"
        t26 = "历史累计登记规模（截至去年" + month + "月" + day + "日）"
        t27 = "历史累计登记只数（截至" + year +"年" + month + "月" + day + "日）"
        t28 = "历史累计登记只数（截至去年" + month + "月" + day + "日）"
        t29 = "存量规模（截至" + year +"年" + month + "月" + day + "日）"
        t30 = "去年存量规模"
        t31 = "去年存量规模（截至" + str(int(year) - 1) +"年" + month + "月" + day + "日）"
        increase = "同比增长"
        title_m_total_lst = [total,t1,t2,increase,t3,t4,increase,t5,t6,increase,t7,t8,increase]
        title_m_trustee_lst = [trustee,t1,t2,increase,t3,t4,increase,t5,t6,increase,t7,t8,increase]
        title_m_deposit_bank_lst = [deposit_bank,t1,t2,increase,t3,t4,increase,t5,t6,increase,t7,t8,increase]

        title_q_total_lst = [total,t9,t10,increase,t11,t12,increase,t13,t14,increase,t15,t16,increase]
        title_q_trustee_lst = [trustee,t9,t10,increase,t11,t12,increase,t13,t14,increase,t15,t16,increase]
        title_q_deposit_bank_lst = [deposit_bank,t9,t10,increase,t11,t12,increase,t13,t14,increase,t15,t16,increase]

        title_y_total_lst = [total,t17,t18,increase,t19,t20,increase,t21,t22,increase,t23,t24,increase]
        title_y_trustee_lst = [trustee,t17,t18,increase,t19,t20,increase,t21,t22,increase,t23,t24,increase]
        title_y_deposit_bank_lst = [deposit_bank,t17,t18,increase,t19,t20,increase,t21,t22,increase,t23,t24,increase]

        title_a_total_lst = [total,t25,t26,increase,t27,t28,increase,t29,t31,increase]
        title_a_trustee_lst = [trustee,t25,t26,increase,t27,t28,increase,t29,t30,increase]
        title_a_deposit_bank_lst = [deposit_bank,t25,t26,increase,t27,t28,increase,t29,t30,increase]

        m_total_lst = self.get_asset_plan_total_reg_release(month_date_start,month_date_end,pre_year_month_start,pre_year_month_end)
        m_trustee = self.get_asset_plan_trustee_reg_release(month_date_start,month_date_end,pre_year_month_start,pre_year_month_end)
        m_deposit_bank = self.get_asset_plan_deposit_bank_reg_release(month_date_start,month_date_end,pre_year_month_start,pre_year_month_end)
        
        q_total_lst = self.get_asset_plan_total_reg_release(quarter_date_start,quarter_date_end,pre_year_quarter_date_start,pre_year_quarter_date_end)
        q_trustee = self.get_asset_plan_trustee_reg_release(quarter_date_start,quarter_date_end,pre_year_quarter_date_start,pre_year_quarter_date_end)
        q_deposit_bank = self.get_asset_plan_deposit_bank_reg_release(quarter_date_start,quarter_date_end,pre_year_quarter_date_start,pre_year_quarter_date_end)
        
        y_total_lst = self.get_asset_plan_total_reg_release(year_date_start,year_date_end,pre_year_date_start,pre_year_date_end)
        y_trustee = self.get_asset_plan_trustee_reg_release(year_date_start,year_date_end,pre_year_date_start,pre_year_date_end)
        y_deposit_bank = self.get_asset_plan_deposit_bank_reg_release(year_date_start,year_date_end,pre_year_date_start,pre_year_date_end)

        a_total_lst = self.get_asset_plan_total_amount(amount_date_init,amount_date_cur,amount_date_init,pre_year_cur_date)
        a_trustee = self.get_asset_plan_trustee_total_amount(amount_date_init,amount_date_cur,amount_date_init,pre_year_cur_date)
        a_deposit_bank = self.get_asset_plan_deposit_bank_amount(amount_date_init,amount_date_cur,amount_date_init,pre_year_cur_date)
        
        xls_data = excel_data()
        xls_data.set_filepath(filepath)
        xls_data.set_month_data(title_m_total_lst,m_total_lst,title_m_trustee_lst,m_trustee,title_m_deposit_bank_lst,m_deposit_bank)
        xls_data.set_quarter_data(title_q_total_lst,q_total_lst,title_q_trustee_lst,q_trustee,title_q_deposit_bank_lst,q_deposit_bank)
        xls_data.set_year_data(title_y_total_lst,y_total_lst,title_y_trustee_lst,y_trustee,title_y_deposit_bank_lst,y_deposit_bank)
        xls_data.set_amount_data(title_a_total_lst,a_total_lst,title_a_trustee_lst,a_trustee,title_a_deposit_bank_lst,a_deposit_bank)
        create_excel_file(xls_data)


    def get_asset_plan_deposit_bank_reg_release(self,date_start,date_end,pre_year_date_start,pre_year_date_end):
        #获得托管银行月度/季度/年度数据
        savelog("获得托管人数据,时间范围:" + date_start + "至" + date_end + "," + pre_year_date_start + "至" + pre_year_date_end)
        #获得托管银行当前年月登记规模
        date_deposit_bank_reg_scope = self.get_deposit_bank_reg_scope(date_start,date_end)
        savelog("托管银行" + date_start + "至" + date_end + "登记规模:" + str(date_deposit_bank_reg_scope))
        #获得托管银行去年当前余额登记规模
        pre_year_date_deposit_bank_reg_scope = self.get_deposit_bank_reg_scope(pre_year_date_start,pre_year_date_end)
        savelog("托管银行" + pre_year_date_start + "至" + pre_year_date_end + "登记规模:" + str(pre_year_date_deposit_bank_reg_scope))

        #获得托管银行当前年月登记只数
        date_desposit_bank_reg_count = self.get_depoist_bank_reg_count(date_start,date_end)
        savelog("托管银行" + date_start + "至" + date_end + "登记只数:" + str(date_desposit_bank_reg_count))
        #获得托管银行去年当前月登记只数
        pre_year_date_deposit_bank_reg_count = self.get_depoist_bank_reg_count(pre_year_date_start,pre_year_date_end)
        savelog("托管银行" + pre_year_date_start + "至" + pre_year_date_end + "登记只数:" + str(pre_year_date_deposit_bank_reg_count))

        #获得托管银行当前年月发行规模
        date_deposit_bank_release_scope = self.get_deposit_bank_release_scope(date_start,date_end)
        savelog("托管银行" + date_start + "至" + date_end + "发行规模:" + str(date_deposit_bank_release_scope))
        #获得托管银行去年当前月发行规模
        pre_year_deposit_bank_release_scope = self.get_deposit_bank_release_scope(pre_year_date_start,pre_year_date_end)
        savelog("托管银行" + pre_year_date_start + "至" + pre_year_date_end + "发行规模:" + str(pre_year_deposit_bank_release_scope))

        #获得托管银行当前年月发行只数
        date_deposit_bank_release_count = self.get_deposit_bank_release_count(date_start,date_end)
        savelog("托管银行" + date_start + "至" + date_end + "发行只数:" + str(date_deposit_bank_release_count))
        #获得托管银行去年当前月发行只数
        pre_year_date_deposit_bank_release_count = self.get_deposit_bank_release_count(pre_year_date_start,pre_year_date_end)
        savelog("托管银行" + pre_year_date_start + "至" + pre_year_date_end + "发行只数:" + str(pre_year_date_deposit_bank_release_count))

        return (date_deposit_bank_reg_scope,pre_year_date_deposit_bank_reg_scope,date_desposit_bank_reg_count,pre_year_date_deposit_bank_reg_count,\
                date_deposit_bank_release_scope,pre_year_deposit_bank_release_scope,date_deposit_bank_release_count,pre_year_date_deposit_bank_release_count)

    def get_asset_plan_trustee_reg_release(self,date_start,date_end,pre_year_date_start,pre_year_date_end):
        #获得受托人情况月度/季度/年度数据
        savelog("获得受托人数据,时间范围:" + date_start + "至" + date_end + "," + pre_year_date_start + "至" + pre_year_date_end)
        #获得受托人情况:当前年,当前月登记规模:
        date_trustee_reg_scope = self.get_trustee_reg_scope(date_start,date_end)
        savelog("受托人" + date_start + "至" + date_end + "登记规模:" + str(date_trustee_reg_scope))
        #获得受托人情况:去年的当前月登记规模
        pre_year_date_trustee_reg_scope = self.get_trustee_reg_scope(pre_year_date_start,pre_year_date_end)
        savelog("受托人" + pre_year_date_start + "至" + pre_year_date_end + "登记规模:" + str(pre_year_date_trustee_reg_scope))
        
        #获得受托人当前年,月登记只数
        date_trustee_reg_count = self.get_trustee_reg_count(date_start,date_end)
        savelog("受托人" + date_start + "至" + date_end + "登记只数:" + str(date_trustee_reg_count))
        #获得受托人情况:去年的当前月登记只数
        pre_year_date_trustee_reg_count = self.get_trustee_reg_count(pre_year_date_start,pre_year_date_end)
        savelog("受托人" + pre_year_date_start + "至" + pre_year_date_end + "登记只数:" + str(pre_year_date_trustee_reg_count))

        #获得受托人当前年,月发行规模
        date_trustee_release_scope = self.get_trustee_release_scope(date_start,date_end)
        savelog("受托人" + date_start + "至" + date_end + "发行规模:" + str(date_trustee_release_scope))
        #获得受托人情况,去年当前月发行规模
        pre_year_trustee_release_scope = self.get_trustee_release_scope(pre_year_date_start,pre_year_date_end)
        savelog("受托人" + pre_year_date_start + "至" + pre_year_date_end + "发行规模:" + str(pre_year_trustee_release_scope))

        # #获得受托人当前年,月发行只数
        date_trustee_release_count = self.get_trustee_release_count(date_start,date_end)
        savelog("受托人" + date_start + "至" + date_end + "发行只数:" + str(date_trustee_release_count))
        #获得受托人情况,去年当前月发行只数
        pre_year_trustee_release_count = self.get_trustee_release_count(pre_year_date_start,pre_year_date_end)
        savelog("受托人" + pre_year_date_start + "至" + pre_year_date_end + "发行只数:" + str(pre_year_trustee_release_count))

        return (date_trustee_reg_scope,pre_year_date_trustee_reg_scope,date_trustee_reg_count,pre_year_date_trustee_reg_count,\
                date_trustee_release_scope,pre_year_trustee_release_scope,date_trustee_release_count,pre_year_trustee_release_count)

    def get_asset_plan_total_reg_release(self,date_start,date_end,pre_year_date_start,pre_year_date_end):
        #获得总体情况月度/季度/年度数据
        savelog("获得总体情况数据,时间范围:" + date_start + "至" + date_end + "," + pre_year_date_start + "至" + pre_year_date_end)
        #获得总体情况:当前年,当前月登记规模:
        date_total_reg_scope = self.get_total_reg_scope(date_start,date_end)
        savelog("总体情况" + date_start + "至" + date_end + "登记规模:" + str(date_total_reg_scope))
        #获得总体情况:去年的当前月登记规模
        pre_year_date_total_reg_scope = self.get_total_reg_scope(pre_year_date_start,pre_year_date_end)
        savelog("总体情况" + pre_year_date_start + "至" + pre_year_date_end + "登记规模:" + str(pre_year_date_total_reg_scope))
        #获得同比增长
        data_increase_0 = "calc"
        data_increase_0 = get_increase_str(pre_year_date_total_reg_scope,date_total_reg_scope)
        
        #获得总体情况:当前年,当前月登记只数:
        date_total_reg_count = self.get_total_reg_count(date_start,date_end)
        savelog("总体情况" + date_start + "至" + date_end + "登记只数:" + str(date_total_reg_count))
        #获得总体情况:去年的当前月登记只数
        pre_year_date_total_reg_count = self.get_total_reg_count(pre_year_date_start,pre_year_date_end)
        savelog("总体情况" + pre_year_date_start + "至" + pre_year_date_end + "登记只数:" + str(pre_year_date_total_reg_count))
        #获得同比增长
        data_increase_1 = "calc"
        data_increase_1 = get_increase_str(pre_year_date_total_reg_count,date_total_reg_count)
        
        #获得总体情况:当前年,月发行规模
        date_total_release_scope = self.get_total_release_scope(date_start,date_end)
        savelog("总体情况" + date_start + "至" + date_end + "发行规模:" + str(date_total_release_scope))
        #获得总体情况:去年发行规模
        pre_year_date_total_release_scope = self.get_total_release_scope(pre_year_date_start,pre_year_date_end)
        savelog("总体情况" + pre_year_date_start + "至" + pre_year_date_end + "发行规模:" + str(pre_year_date_total_release_scope))
        #获得同比增长
        data_increase_2 = "calc"
        data_increase_2 = get_increase_str(pre_year_date_total_release_scope,date_total_release_scope)
        
        #获得总体情况:当前年,月发行只数
        date_total_release_count = self.get_total_release_count(date_start,date_end)
        savelog("总体情况" + date_start + "至" + date_end + "发行只数:" + str(date_total_release_count))
        #获得总体情况：去年发行只数
        pre_year_date_total_release_count = self.get_total_release_count(pre_year_date_start,pre_year_date_end)
        savelog("总体情况" + pre_year_date_start + "至" + pre_year_date_end + "发行只数:" + str(pre_year_date_total_release_count))
        #获得同比增长
        data_increase_3 = "calc"
        data_increase_3 = get_increase_str(pre_year_date_total_release_count,date_total_release_count)
        
        reslist = ["",date_total_reg_scope,pre_year_date_total_reg_scope,data_increase_0,\
                    date_total_reg_count,pre_year_date_total_reg_count,data_increase_1,\
                    date_total_release_scope,pre_year_date_total_release_scope,data_increase_2,\
                    date_total_release_count,pre_year_date_total_release_count,data_increase_3]
        return reslist

    def get_asset_plan_total_amount(self,date_start,date_end,pre_year_date_start,pre_year_date_end):
        #获得总体情况历史累计数据
        savelog("获得历史累计总体情况数据,时间范围:" + date_start + "至" + date_end + "," + pre_year_date_start + "至" + pre_year_date_end)
        #获得总体情况:历史累计登记规模:
        date_total_reg_scope = self.get_total_reg_scope(date_start,date_end)
        savelog("历史累计,总体情况" + date_start + "至" + date_end + "登记规模:" + str(date_total_reg_scope))
        #获得总体情况:历史累计登记规模（截至去年x月x日）
        pre_year_date_total_reg_scope = self.get_total_reg_scope(pre_year_date_start,pre_year_date_end)
        savelog("历史累计,总体情况" + pre_year_date_start + "至" + pre_year_date_end + "登记规模:" + str(pre_year_date_total_reg_scope))
        #获得同比增长
        data_increase_0 = "calc"
        data_increase_0 = get_increase_str(pre_year_date_total_reg_scope,date_total_reg_scope)
        # savelog("总体情况历史累计登记规模同比增长:" + str(data_increase_0))

        #获得总体情况:历史累计登记只数（截至xx年x月x日）:
        date_total_reg_count = self.get_total_reg_count(date_start,date_end)
        savelog("历史累计,总体情况" + date_start + "至" + date_end + "登记只数:" + str(date_total_reg_count))
        #获得总体情况:历史累计登记只数（截至去年x月x日）
        pre_year_date_total_reg_count = self.get_total_reg_count(pre_year_date_start,pre_year_date_end)
        savelog("历史累计,总体情况" + pre_year_date_start + "至" + pre_year_date_end + "登记只数:" + str(pre_year_date_total_reg_count))
        #获得同比增长
        data_increase_1 = "calc"
        data_increase_1 = get_increase_str(pre_year_date_total_reg_count,date_total_reg_count)
        # savelog("总体情况历史累计登记只数同比增长:" + str(data_increase_1))

        #获得总体情况:存量规模（截至xx年x月x日）
        date_total_release_scope = self.get_amount_total_scope(date_end)
        savelog("历史累计,总体情况" + date_start + "至" + date_end + "存量规模:" + str(date_total_release_scope))
        #获得总体情况:去年存量规模（截至xx年x月x日）
        pre_year_date_total_release_scope = self.get_amount_total_scope(pre_year_date_end)
        savelog("历史累计,总体情况" + pre_year_date_start + "至" + pre_year_date_end + "存量规模:" + str(pre_year_date_total_release_scope))
        #获得同比增长
        data_increase_2 = "calc"
        data_increase_2 = get_increase_str(pre_year_date_total_release_scope,date_total_release_scope)
        # savelog("总体情况历史累计存量规模同比增长:" + str(data_increase_2))

        reslist = ["",date_total_reg_scope,pre_year_date_total_reg_scope,data_increase_0,\
                    date_total_reg_count,pre_year_date_total_reg_count,data_increase_1,\
                    date_total_release_scope,pre_year_date_total_release_scope,data_increase_2]
        return reslist
    
    def get_asset_plan_trustee_total_amount(self,date_start,date_end,pre_year_date_start,pre_year_date_end):
        #获得受托人情况历史累计数据
        savelog("获得历史累计受托人数据,时间范围:" + date_start + "至" + date_end + "," + pre_year_date_start + "至" + pre_year_date_end)
        #获得受托人情况:历史累计登记规模（截至xx年x月x日）:
        date_trustee_reg_scope = self.get_trustee_reg_scope(date_start,date_end)
        savelog("历史累计,受托人" + date_start + "至" + date_end + "登记规模:" + str(date_trustee_reg_scope))
        #获得受托人情况:历史累计登记规模（截至去年x月x日）
        pre_year_date_trustee_reg_scope = self.get_trustee_reg_scope(pre_year_date_start,pre_year_date_end)
        savelog("历史累计,受托人" + pre_year_date_start + "至" + pre_year_date_end + "登记规模:" + str(pre_year_date_trustee_reg_scope))
        
        #获得受托人历史累计登记只数（截至xx年x月x日）
        date_trustee_reg_count = self.get_trustee_reg_count(date_start,date_end)
        savelog("历史累计,受托人" + date_start + "至" + date_end + "登记只数:" + str(date_trustee_reg_count))
        #获得受托人情况:历史累计登记只数（截至去年x月x日）
        pre_year_date_trustee_reg_count = self.get_trustee_reg_count(pre_year_date_start,pre_year_date_end)
        savelog("历史累计,受托人" + pre_year_date_start + "至" + pre_year_date_end + "登记只数:" + str(pre_year_date_trustee_reg_count))

        #获得受托人存量规模（截至xx年x月x日）
        date_trustee_release_scope = self.get_amount_trustee_scope(date_end)
        savelog("历史累计,受托人" + date_start + "至" + date_end + "存量规模:" + str(date_trustee_release_scope))
        #获得受托人情况,去年存量规模（截至xx年x月x日）
        pre_year_trustee_release_scope = self.get_amount_trustee_scope(pre_year_date_end)
        savelog("历史累计,受托人" + pre_year_date_start + "至" + pre_year_date_end + "存量规模:" + str(pre_year_trustee_release_scope))

        return (date_trustee_reg_scope,pre_year_date_trustee_reg_scope,date_trustee_reg_count,pre_year_date_trustee_reg_count,date_trustee_release_scope,pre_year_trustee_release_scope)
    
    def get_asset_plan_deposit_bank_amount(self,date_start,date_end,pre_year_date_start,pre_year_date_end):
        #获得托管银行历史累计
        savelog("获得历史累计托管银行数据,时间范围:" + date_start + "至" + date_end + "," + pre_year_date_start + "至" + pre_year_date_end)
        #获得托管银行历史累计登记规模（截至xx年x月x日）
        date_deposit_bank_reg_scope = self.get_deposit_bank_reg_scope(date_start,date_end)
        savelog("历史累计,托管银行" + date_start + "至" + date_end + "登记规模:" + str(date_deposit_bank_reg_scope))
        #获得托管银行历史累计登记规模（截至去年x月x日）
        pre_year_date_deposit_bank_reg_scope = self.get_deposit_bank_reg_scope(pre_year_date_start,pre_year_date_end)
        savelog("历史累计,托管银行" + pre_year_date_start + "至" + pre_year_date_end + "登记规模:" + str(pre_year_date_deposit_bank_reg_scope))

        #获得托管银行历史累计登记只数（截至xx年x月x日）
        date_desposit_bank_reg_count = self.get_depoist_bank_reg_count(date_start,date_end)
        savelog("历史累计,托管银行" + date_start + "至" + date_end + "登记只数:" + str(date_desposit_bank_reg_count))
        #获得托管银行历史累计登记只数（截至去年x月x日）
        pre_year_date_deposit_bank_reg_count = self.get_depoist_bank_reg_count(pre_year_date_start,pre_year_date_end)
        savelog("历史累计,托管银行" + pre_year_date_start + "至" + pre_year_date_end + "登记只数:" + str(pre_year_date_deposit_bank_reg_count))

        #获得托管银行存量规模（截至xx年x月x日）
        date_deposit_bank_release_scope = self.get_amount_deposit_bank_scope(date_end)
        savelog("历史累计,托管银行" + date_start + "至" + date_end + "存量规模:" + str(date_deposit_bank_release_scope))
        #获得托管银行去年存量规模（截至xx年x月x日）
        pre_year_deposit_bank_release_scope = self.get_amount_deposit_bank_scope(pre_year_date_end)
        savelog("历史累计,托管银行" + pre_year_date_start + "至" + pre_year_date_end + "存量规模:" + str(pre_year_deposit_bank_release_scope))

        return (date_deposit_bank_reg_scope,pre_year_date_deposit_bank_reg_scope,date_desposit_bank_reg_count,pre_year_date_deposit_bank_reg_count,date_deposit_bank_release_scope,pre_year_deposit_bank_release_scope)
    
    def get_caption_from_dict(self,val,keyno):
        #从fundcrm.tdictionary获得基础资产类型
        if val is None:
            return ""
        valist = val.split(",")
        reslst = []
        for item in valist:
            sql = '''
                SELECT C_CAPTION FROM fundcrm.tdictionary WHERE  L_KEYNO = key_no AND C_KEYVALUE='value' ORDER BY C_KEYVALUE
                '''
            sql = sql.replace('value',item).replace('key_no',keyno)
            savelog("查询基础资产类型:" + sql)
            dataresult = self.cursor.execute(sql)
            data = dataresult.fetchall()
            savelog("查询结果:"+str(data),False)
            v = str(item)
            if data[0][0] is not None:
                v = str(data[0][0])
            reslst.append(v)
        valstr = str(reslst)
        valstr = valstr.replace("'","").replace('[','').replace(']','')
        return valstr

    def get_report_date(self,fundcode):
        #报送日期查询
        sql = '''
            SELECT
                END_
            FROM
                (
                    SELECT
                        T.END_
                    FROM
                        fundcrm.jbpm4_ext_hist_task E,
                        fundcrm.jbpm4_hist_task T,
                        fundcrm.jbpm4_hist_procinst P
                    WHERE
                        E.dbid_ = T.dbid_
                    AND T.procinst_ = P.dbid_
                    AND E.biz_flow_uuid_ = (
                        SELECT
                            f.c_fundcode
                        FROM
                            fundcrm.tfundinfo f
                        WHERE
                            f.c_Tafundcode = 'tafundcode'
                    )
                    AND P.proc_exe_name_ = '产品登记流程'
                    AND E.EXTFIELD1 = '机构复核通过'
                    AND T.outcome_ = 'pass'
                    ORDER BY
                        T.END_
                )
            WHERE
                ROWNUM = 1
        '''
        sql = sql.replace('tafundcode',fundcode)
        savelog("报送日期查询SQL:" + sql)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:" + str(data),False)
        if data is not None and len(data) > 0 and data[0][0] is not None:
            return switch_reg_date(str(data[0][0]))
        else:
            return ""

    def check_workday(self,timestr):
        #判断当前时间是否是工作日
        sql = "SELECT L_WORKFLAG FROM fundcrm.TOPENDAY WHERE to_char(D_DATE , 'yyyy-mm-dd') = 'datetimestr'"
        tstr = ""
        if timestr.find(' ') != -1:
            tstr = timestr.split(' ')[0]
        else:
            tstr = timestr
        sql = sql.replace('datetimestr',tstr)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog(sql + " 查询结果:" + str(data),False)
        if data is None or len(data) == 0 or data[0][0] is None or str(data[0][0]) == "None":
            return False
        if int(data[0][0]) == 1:
            return True
        else:
            return False

    def get_work_day_count(self,daylst):
        count = 0
        for item in daylst:
            if self.check_workday(str(item)):
                count = count + 1
        return count

    def get_reg_month_asset(self,start_date,end_date):
        #获得中保登产品登记情况的资产支持计划
        sql = '''
            SELECT
                a.C_TAFUNDCODE,
                --产品名称
                a.C_FUNDNAME,
                --管理人
                d.mem_name,
                --登记规模
                b.F_PRDREGBALA,
                --基础资产类型 3734
                c.C_INFRACLASS,
                --登记日期
                e.d_preregisterdate,
                --登记状态
                '予以登记' 
            FROM
                fundcrm.tfundinfo a,
                fundcrm.TPDTBASICINFO b,
                fundcrm.TPDTINFRAINFO c,
                member.member_info d,
                fundcrm.tpdtoperateinfo e
            WHERE
                a.c_fundcode = b.c_fundcode(+)
            AND a.c_fundcode = c.c_fundcode(+)
            AND b.c_mgrorgno = d.mem_code(+)
            AND a.c_fundcode = e.c_fundcode(+)
            AND a.C_PDTTYPE = 1 
            AND a.L_PDTTEMPLETID = 8 
            AND a.C_AUDITFLAG in (3,4,5,6,7,12,13,15,16,20,21,22,25,29,55,56,57,58)
            ORDER BY e.D_PREREGISTERDATE DESC NULLS LAST
        '''
        savelog("资产支持计划登记情况:" + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(len(data)) + "条数据",False)
        resmap = {}
        line = 0
        for item in data:
            caplst = self.get_caption_from_dict(item[4],"3734")
            report_date = self.get_report_date(str(item[0]))
            mem_name = get_default_value(item[2])
            regscope = ""
            if item[3] is not None and str_is_float(item[3]):
                regscope = float(item[3]) / 100000000.0

            prdregbala = get_default_value(regscope,"0")
            register_date = get_default_value(get_truncate_date(str(item[5])))
            days_between = list_dates_between(str(report_date),str(register_date))
            working_days = ""
            if len(days_between) > 0:
                working_days = self.get_work_day_count(days_between)

            resmap[str(line)] = {
                "C_TAFUNDCODE":str(item[0]),
                "C_FUNDNAME":str(item[1]),#产品名称
                "mem_name":str(mem_name),#管理人
                "F_PRDREGBALA":str(prdregbala),#登记规模
                "C_INFRACLASS":str(caplst),#基础资产类型
                "d_preregisterdate":str(register_date),#登记日期
                "C_AUDITFLAG":str(item[6]),#登记状态
                "reportdate":str(report_date), #报送日期
                "days_between":str(working_days) #登记耗时
            }
            # strline = str(item[0]) + " " + str(item[1]) + " " + mem_name + " " + str(prdregbala) + " " + str(item[4]) + "(" + str(caplst) + ") " + \
            #             str(register_date) + " " + str(item[6] + " " + str(report_date) + " days_between:" + str(days_between) + " working_days:" + str(working_days))
            # savelog(strline)
            line = line + 1

        return resmap

    def get_reg_group_asset(self):
        # 获得“予以登记”组合资管产品明细
        sql = '''
            SELECT
                a.C_TAFUNDCODE,
                    a.C_FUNDNAME,--产品名称
                    c.C_PRDTYPE,--产品类型    数据字典编号 5024
                    a.C_OPERATEWAY,--运作方式 数据字典编号  5132
                    b.C_REGISTERCLASSIY,--登记方式 数据字典编号 9031
                    '予以登记'--流程进度
                FROM
                    fundcrm.tfundinfo a,
                    fundcrm.TPDTEXTENDINFO b,
                    fundcrm.TPDTBASICINFO c
                WHERE
                    a.c_fundcode = b.c_fundcode(+)
                    AND a.c_fundcode = c.c_fundcode(+)
                    AND a.C_AUDITFLAG in(32,45,46,47,50,51,52,53,54)
                    AND a.L_PDTTEMPLETID in(10,17)
        '''
        savelog("获得予以登记组合资管产品明细:" + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(len(data)) + "条数据",False)
        resmap = {}
        line = 0
        for item in data:
            savelog(str(line) + " " + str(item))
            fundname = get_default_value(item[0])
            prdtype = self.get_caption_from_dict(item[2],"5024")
            operateway = self.get_caption_from_dict(item[3],"5132")
            registerclassiy = self.get_caption_from_dict(item[4],"9031")
            reg_progress = get_default_value(item[5])
            resmap[str(line)] = {
                "C_FUNDNAME":str(fundname),
                "C_PRDTYPE":str(prdtype),
                "C_OPERATEWAY":str(operateway),
                "C_REGISTERCLASSIY":str(registerclassiy),
                "C_REG_PROGRESS":str(reg_progress)
            }
            line = line + 1

        return resmap

    def get_product_type(self,month_lst):
        #获得"产品类型"统计结果
        mid = 1
        resultmap = {"1":None,"2":None,"3":None,"4":None,"5":None,"6":None,"7":None,"8":None,"9":None,"10":None,"11":None,"12":None}
        for i in month_lst:
            start_date = i[0]
            end_date = i[1]
            sql = '''
                SELECT
                    c.C_PRDTYPE,COUNT(c.C_PRDTYPE) 
                    FROM
                        fundcrm.tfundinfo a,
                        fundcrm.TPDTEXTENDINFO b,
                        fundcrm.TPDTBASICINFO c,
                        fundcrm.tpdtoperateinfo d
                    WHERE
                        a.c_fundcode = b.c_fundcode(+)
                        AND a.c_fundcode = c.c_fundcode(+)
                        AND a.c_fundcode = d.c_fundcode(+)
                        AND a.C_AUDITFLAG in(32,45,46,47,50,51,52,53,54)
                        AND a.L_PDTTEMPLETID in(10,17)
                        AND TO_CHAR(d.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(d.d_preregisterdate,'yyyy-mm-dd')<='enddate'
                GROUP BY c.C_PRDTYPE
            '''
            sql = sql.replace("startdate",start_date).replace("enddate",end_date)
            savelog("产品类型(只)统计结果,日期范围:" + start_date + "至" + end_date + sql,False)
            dataresult = self.cursor.execute(sql)
            data = dataresult.fetchall()
            savelog("查询结果:"+str(data),False)
            regmap = {"0":"0","1":"0","2":"0","9":"0"}
            # 0:固定收益类
            # 1:权益类
            # 2:商品及金融衍生品类
            # 9:混合类
            for item in data:
                k = ""
                v = ""
                if item[0] is not None:
                    k = item[0]
                if item[1] is not None:
                    v = item[1]

                regmap[str(k)] = str(v)

            resultmap[str(mid)] = regmap
            mid = mid + 1

        return resultmap

    def get_working_type(self,month_lst):
        #获得运作方式统计结果
        mid = 1
        resultmap = {"1":None,"2":None}
        for i in month_lst:
            start_date = i[0]
            end_date = i[1]
            sql = '''
                SELECT
                    a.C_OPERATEWAY,COUNT(a.C_OPERATEWAY) 
                    FROM
                        fundcrm.tfundinfo a,
                        fundcrm.TPDTEXTENDINFO b,
                        fundcrm.TPDTBASICINFO c,
                        fundcrm.tpdtoperateinfo d
                    WHERE
                        a.c_fundcode = b.c_fundcode(+)
                        AND a.c_fundcode = c.c_fundcode(+)
                        AND a.c_fundcode = d.c_fundcode(+)
                        AND a.C_AUDITFLAG in(32,45,46,47,50,51,52,53,54)
                        AND a.L_PDTTEMPLETID in(10,17)
                        AND TO_CHAR(d.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(d.d_preregisterdate,'yyyy-mm-dd')<='enddate'
                GROUP BY a.C_OPERATEWAY
            '''
            sql = sql.replace("startdate",start_date).replace("enddate",end_date)
            savelog("运作方式（只）统计结果,日期范围:" + start_date + "至" + end_date + sql,False)
            dataresult = self.cursor.execute(sql)
            data = dataresult.fetchall()
            savelog("查询结果:"+str(data),False)
            workingmap = {"1":"0","2":"0"}
            # 1:开放式
            # 2:封闭式
            for item in data:
                k = ""
                v = ""
                if item[0] is not None:
                    k = item[0]
                if item[1] is not None:
                    v = item[1]

                workingmap[str(k)] = str(v)

            resultmap[str(mid)] = workingmap
            mid = mid + 1

        return resultmap

    def get_manager_product_type(self,start_date,end_date):
        #按照 产品管理人、产品类型 进行统计
        sql = '''
                SELECT
                    e.mem_name,c.C_PRDTYPE,COUNT(c.C_PRDTYPE) 
                    FROM
                        fundcrm.tfundinfo a,
                        fundcrm.TPDTEXTENDINFO b,
                        fundcrm.TPDTBASICINFO c,
                        fundcrm.tpdtoperateinfo d,
                        member.member_info e
                    WHERE
                        a.c_fundcode = b.c_fundcode(+)
                        AND a.c_fundcode = c.c_fundcode(+)
                        and c.c_mgrorgno = e.mem_code(+)
                        AND a.c_fundcode = d.c_fundcode(+)
                        AND a.C_AUDITFLAG in(32,45,46,47,50,51,52,53,54)
                        AND a.L_PDTTEMPLETID in(10,17)
                        AND TO_CHAR(d.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(d.d_preregisterdate,'yyyy-mm-dd')<='enddate'
                GROUP BY e.mem_name,c.C_PRDTYPE
                ORDER BY e.mem_name
            '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("产品管理人、产品类型 进行统计,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(data),False)
        resultmap = {}
        for item in data:
            key = item[0]
            vmap = {}
            vmap[str(item[1])] = str(item[2])
            if key in resultmap:
                v = resultmap[key]
                v.update(vmap)
                resultmap[key] = v
            else:
                resultmap[key] = vmap

        return resultmap

    def get_manager_working_type(self,start_date,end_date):
        #产品管理人、运作方式 进行统计
        sql = '''
            SELECT
                e.mem_name,a.C_OPERATEWAY,COUNT(a.C_OPERATEWAY) 
                FROM
                    fundcrm.tfundinfo a,
                    fundcrm.TPDTEXTENDINFO b,
                    fundcrm.TPDTBASICINFO c,
                    fundcrm.tpdtoperateinfo d,
                    member.member_info e
                WHERE
                    a.c_fundcode = b.c_fundcode(+)
                    AND a.c_fundcode = c.c_fundcode(+)
                    and c.c_mgrorgno = e.mem_code(+)
                    AND a.c_fundcode = d.c_fundcode(+)
                    AND a.C_AUDITFLAG in(32,45,46,47,50,51,52,53,54)
                    AND a.L_PDTTEMPLETID in(10,17)
                    AND TO_CHAR(d.d_preregisterdate,'yyyy-mm-dd')>='startdate' AND  TO_CHAR(d.d_preregisterdate,'yyyy-mm-dd')<='enddate'
            GROUP BY e.mem_name,a.C_OPERATEWAY
            '''
        sql = sql.replace("startdate",start_date).replace("enddate",end_date)
        savelog("产品管理人、运作方式 进行统计,日期范围:" + start_date + "至" + end_date + sql,False)
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        savelog("查询结果:"+str(data),False)
        resultmap = {}
        for item in data:
            key = item[0]
            vmap = {}
            vmap[str(item[1])] = str(item[2])
            if key in resultmap:
                v = resultmap[key]
                v.update(vmap)
                resultmap[key] = v
            else:
                resultmap[key] = vmap

        return resultmap


    def create_asset_plan_reg_month_file(self,filepath,y="",m="",d=""):
        #创建《附件二：中保登产品登记情况月报表.xlsx》
        savelog("create_asset_plan_reg_month_file filepath:" + filepath)
        datemap = get_date_info_map(y,m,d)
        month_date_start = datemap["month_date_start"]
        month_date_end = datemap["month_date_end"]
        asset_map = self.get_reg_month_asset(month_date_start,month_date_end)
        reg_group_asset_map = self.get_reg_group_asset()
        reg_month_map = self.get_product_type(datemap["current_year_months_scope"])
        working_type_map = self.get_working_type(datemap["current_year_months_scope"])
        manager_product_type_map = self.get_manager_product_type(datemap["year_date_start"],datemap["year_date_end"])
        manager_working_type_map = self.get_manager_working_type(datemap["year_date_start"],datemap["year_date_end"])

        asset_support_month = excel_data_month()
        asset_support_month.set_filepath(filepath)
        asset_support_month.set_asset_support_plan_data(asset_map)
        asset_support_month.set_reg_group_asset_data(reg_group_asset_map)
        asset_support_month.set_product_reg_type_data(reg_month_map)
        asset_support_month.set_working_type_date(working_type_map)
        asset_support_month.set_manager_product_type_data(manager_product_type_map)
        asset_support_month.set_manager_working_type_data(manager_working_type_map)
        create_reg_month_excel_file(asset_support_month)
    
    def getTempleName(self,templetId):
        if templetId == '8':
            return '资产支持计划'
        elif templetId == '6':
            return '债权'
        elif templetId == '7':
            return '股权'
        else:
            return templetId

    def sendToDestFtp(self,localDir,date,fileDestHost,fileDestPort,fileDestUsername,fileDestPassword,fileDestRemotePath):
        fileDestSftp = self.getFtp(fileDestHost,fileDestPort,fileDestUsername,fileDestPassword)

        fileDestSftp.stat(fileDestRemotePath)
        stdin, stdout, stderr = fileDestSftp.exec_command('ls DIR')
        if stdout.readline() != '':
            print("exist")
        else:
            print("not exist")

        fileDestSftp.chdir(fileDestRemotePath)
        fileDestSftp.mkdir(date);
        fileDestSftp.put()

    def get_filelist(self,dir, Filelist):
        newDir = dir
        if os.path.isfile(dir):
            Filelist.append(dir)
        elif os.path.isdir(dir):
            for s in os.listdir(dir):
                newDir = os.path.join(dir, s)
                self.get_filelist(newDir, Filelist)
        return Filelist

    # 删除
    def del_file(self,path):
        if os.path.exists(path):
            ls = os.listdir(path)
            for i in ls:
                c_path = os.path.join(path, i)
                if os.path.isdir(c_path):
                    self.del_file(c_path)
                else:
                    os.remove(c_path)
            os.rmdir(path)

