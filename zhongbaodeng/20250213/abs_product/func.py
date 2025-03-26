from datetime import date,datetime
import datetime
import time;
import calendar
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side, Font

class excel_data(object):
    filepath = ""
    title_month_total_lst = []
    month_total_lst = []
    title_month_trustee_lst = []
    m_trustee = None
    title_month_deposit_bank_lst = []
    m_deposit_bank = None

    title_quarter_total_lst = []
    quarter_total_lst = []
    title_quarter_trustee_lst = []
    q_trustee = None
    title_quarter_deposit_bank_lst = []
    q_deposit_bank = None

    title_year_total_lst = []
    year_total_lst = []
    title_year_trustee_lst = []
    y_trustee = None
    title_year_deposit_bank_lst = []
    y_deposit_bank = None

    title_amount_total_lst = []
    amount_total_lst = []
    title_amount_trustee_lst = []
    a_trustee = None
    title_amount_deposit_bank_lst = []
    a_deposit_bank = None

    def get_filepath(self):
        return self.filepath
    
    def set_filepath(self,f):
        self.filepath = f

    def set_month_data(self,t_m_total_lst,m_total_lst,t_m_trustee_lst,m_trustee_tup,t_m_deposit_bank_lst,m_deposit_bank_tup):
        #设置月度数据
        self.title_month_total_lst = t_m_total_lst
        self.month_total_lst = m_total_lst
        self.title_month_trustee_lst = t_m_trustee_lst
        self.m_trustee = m_trustee_tup
        self.title_month_deposit_bank_lst = t_m_deposit_bank_lst
        self.m_deposit_bank = m_deposit_bank_tup
    
    def set_quarter_data(self,t_q_total_lst,q_total_lst,t_q_trustee_lst,q_trustee_tup,t_q_deposit_bank_lst,q_deposit_bank_tup):
        #设置季度数据
        self.title_quarter_total_lst = t_q_total_lst
        self.quarter_total_lst = q_total_lst
        self.title_quarter_trustee_lst = t_q_trustee_lst
        self.q_trustee = q_trustee_tup
        self.title_quarter_deposit_bank_lst = t_q_deposit_bank_lst
        self.q_deposit_bank = q_deposit_bank_tup

    def set_year_data(self,t_y_total_lst,y_total_lst,t_y_trustee_lst,y_trustee_tup,t_y_deposit_bank_lst,y_deposit_bank_tup):
        #设置年度数据
        self.title_year_total_lst = t_y_total_lst
        self.year_total_lst = y_total_lst
        self.title_year_trustee_lst = t_y_trustee_lst
        self.y_trustee = y_trustee_tup
        self.title_year_deposit_bank_lst = t_y_deposit_bank_lst
        self.y_deposit_bank = y_deposit_bank_tup

    def set_amount_data(self,t_a_total_lst,a_total_lst,t_a_trustee_lst,a_trustee_tup,t_a_deposit_bank_lst,a_deposit_bank_tup):
        #设置历史累计数据
        self.title_amount_total_lst = t_a_total_lst
        self.amount_total_lst = a_total_lst
        self.title_amount_trustee_lst = t_a_trustee_lst
        self.a_trustee = a_trustee_tup
        self.title_amount_deposit_bank_lst = t_a_deposit_bank_lst
        self.a_deposit_bank = a_deposit_bank_tup


class excel_data_month(object):
    filepath = ""
    title1 = "资产支持计划"
    title2 = "组合产品"
    txt1 = "资产支持计划登记情况"
    txt2 = "\"予以登记\"组合资管产品明细"
    txt3 = "\"予以登记\"组合资管产品月度统计"
    txt4 = "产品类型（只）"
    txt5 = "运作方式（只）"
    lst1 = ["序号","产品名称","管理人","登记规模（亿元）","基础资产类型","登记日期","登记耗时（工作日）","报送日期","登记状态"]
    lst2 = ["序号","产品名称","产品类型","运作方式","登记方式","流程进度"]
    lst3 = ["固定收益类","权益类","商品及金融衍生品类","混合类","开放式","封闭式"]
    alphabet = ['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    fill_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid") 
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border_style = Border(
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'),
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000')
        )
    font1 = Font(bold=True, size=12)
    #资产支持计划登记情况
    asset_support_plan_map = None

    # “予以登记”组合资管产品明细
    reg_group_asset_map = None

    #产品类型（只）
    product_reg_type_map = None

    #运作方式（只）
    working_type_map = None

    #产品管理人产品类型统计数据
    manager_product_type = None

    #产品管理人运作方式统计数据
    manager_working_type = None

    def get_filepath(self):
        return self.filepath
    
    def set_filepath(self,f):
        self.filepath = f

    def set_asset_support_plan_data(self,asset_support_plan_map):
        self.asset_support_plan_map = asset_support_plan_map

    def set_reg_group_asset_data(self,data_map):
        self.reg_group_asset_map = data_map

    def set_product_reg_type_data(self,data_map):
        self.product_reg_type_map = data_map

    def set_working_type_date(self,data_map):
        self.working_type_map = data_map

    def set_manager_product_type_data(self,data_map):
        self.manager_product_type = data_map

    def set_manager_working_type_data(self,data_map):
        self.manager_working_type = data_map

class asset_reg_data_month(object):
    filepath = ""
    title1 = "资产支持计划"
    title2 = "组合产品"
    txt1 = "资产支持计划登记情况"
    txt2 = "\"予以登记\"组合资管产品明细"
    txt3 = "\"予以登记\"组合资管产品月度统计"
    txt4 = "产品类型（只）"
    txt5 = "运作方式（只）"
    lst1 = ["序号","产品名称","管理人","登记规模（亿元）","基础资产类型","登记日期","登记耗时（工作日）","报送日期","登记状态"]
    lst2 = ["序号","产品名称","产品类型","运作方式","登记方式","流程进度"]
    lst3 = ["固定收益类","权益类","商品及金融衍生品类","混合类","开放式","封闭式"]
    alphabet = ['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    fill_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid") 
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border_style = Border(
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'),
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000')
        )
    font1 = Font(bold=True, size=12)
    wb = None
    ws1 = None
    ws2 = None

    def __init__(self,f):
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.filepath = f

    def save_sheet(self):
        self.wb.active = self.ws1
        self.wb.save(self.filepath)
        print(self.filepath + " 文件已生成！")
        
    def set_ws1_sheet(self):
        self.ws1.title = self.title1
        self.ws1.merge_cells('A1:I1')
        self.ws1['A1'] = self.txt1
        self.ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws1['A1'].fill = self.fill_grey
        self.ws1['A1'].font = self.font1
        
        self.set_excel_sheet_row_text(2,self.ws1,self.lst1,1,True,self.font1)
        self.set_merge_cells_border_row(self.ws1,"A1","I1")

    def set_ws2_sheet(self):
        self.ws2 = self.wb.create_sheet(self.title2)
        self.ws2.merge_cells('A1:F1')
        self.ws2['A1'] = self.txt2
        self.ws2['A1'].font = self.font1
        self.ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        self.set_merge_cells_border_row(self.ws2,"A1","F1")

        self.set_excel_sheet_row_text(2,self.ws2,self.lst2,1,False,self.font1)

        self.ws2.merge_cells('A6:G6')
        self.ws2['A6'] = self.txt3
        self.ws2['A6'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws2['A6'].font = self.font1
        self.set_merge_cells_border_row(self.ws2,"A6","G6")
        
        self.ws2.merge_cells('A7:A8')
        self.ws2['A7'] = "月份"
        self.ws2['A7'].font = self.font1
        self.ws2['A7'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws2['A7'].border = self.border_style

        self.ws2.merge_cells('B7:E7')
        self.ws2['B7'] = self.txt4
        self.ws2['B7'].font = self.font1
        self.ws2['B7'].alignment = Alignment(horizontal='center', vertical='center')
        self.set_merge_cells_border_row(self.ws2,"B7","E7")
        
        self.ws2.merge_cells('F7:G7')
        self.ws2['F7'] = self.txt5
        self.ws2['F7'].font = self.font1
        self.ws2['F7'].alignment = Alignment(horizontal='center', vertical='center')
        self.set_merge_cells_border_row(self.ws2,"F7","G7")
        self.set_excel_sheet_row_text(8,self.ws2,self.lst3,2,False,self.font1)
        
        self.set_month(self.ws2,9,1)

        self.ws2.merge_cells('A24:I24')
        self.ws2['A24'] = self.txt3
        self.ws2['A24'].font = self.font1
        self.ws2['A24'].alignment = Alignment(horizontal='center', vertical='center')
        self.set_merge_cells_border_row(self.ws2,"A24","I24")

        self.ws2.merge_cells('A25:A26')
        self.ws2['A25'] = '序号'
        self.ws2['A25'].font = self.font1
        self.ws2['A25'].alignment = Alignment(horizontal='center', vertical='center')
        self.set_merge_cells_border_col(self.ws2,"A25","A26")

        self.ws2.merge_cells('B25:B26')
        self.ws2['B25'] = '管理人'
        self.ws2['B25'].font = self.font1
        self.ws2['B25'].alignment = Alignment(horizontal='center', vertical='center')
        self.set_merge_cells_border_col(self.ws2,"B25","B26")

        self.ws2.merge_cells('C25:F25')
        self.ws2['C25'] = '产品类型'
        self.ws2['C25'].font = self.font1
        self.ws2['C25'].alignment = Alignment(horizontal='center', vertical='center')
        self.set_merge_cells_border_row(self.ws2,"C25","F25")

        self.ws2.merge_cells('G25:H25')
        self.ws2['G25'] = '运作方式'
        self.ws2['G25'].font = self.font1
        self.ws2['G25'].alignment = Alignment(horizontal='center', vertical='center')
        self.set_merge_cells_border_row(self.ws2,"G25","H25")

        self.ws2.merge_cells('I25:I26')
        self.ws2['I25'] = '合计'
        self.ws2['I25'].font = self.font1
        self.ws2['I25'].alignment = Alignment(horizontal='center', vertical='center')
        self.set_merge_cells_border_row(self.ws2,"G25","H25")
        self.set_merge_cells_border_col(self.ws2,"I25","I26")

        self.set_excel_sheet_row_text(26,self.ws2,self.lst3,3,False,self.font1)

    def set_merge_cells_border_col(self,ws,bstart,bend):
        cell = bstart[0]
        start = int(bstart[1:])
        end = int(bend[1:])
        while start <= end:
            ws[cell + str(start)].border = self.border_style
            start = start + 1

    def set_merge_cells_border_row(self,ws,bstart,bend):
        #设置合并后的单元格边框
        start = bstart[0]
        end = bend[0]
        row = str(int(bstart[1:]))
        i = 1
        bstart = False
        while i < len(self.alphabet):
            if start == self.alphabet[i]:
                bstart = True
            if bstart and end == self.alphabet[i]:
                bstart = False
                ws[self.alphabet[i] + row].border = self.border_style
            
            if bstart == True:
                ws[self.alphabet[i] + row].border = self.border_style

            i = i + 1

    def set_excel_sheet_row_text(self,line,ws,text_list,id,fill_color = True,font = None):
        fill_grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for item in text_list:
            ws.cell(row=line,column=id).value = item
            cell_id = self.alphabet[id]
            ws.column_dimensions[cell_id].width = 15
            if fill_color:
                ws[cell_id + str(line)].fill = fill_grey
            if font is not None:
                ws[cell_id + str(line)].font = font
            ws[cell_id + str(line)].border = self.border_style
            if len(str(item)) <= 6:
                ws[cell_id + str(line)].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[cell_id + str(line)].alignment = Alignment(wrap_text=True)

            id = id + 1

    def set_month(self,ws,line,id):
        for i in range(12):
            ws.cell(row=line,column=id).value = str(i + 1) + "月"
            cell_id = self.alphabet[id]
            ws[cell_id + str(line)].border = self.border_style
            ws[cell_id + str(line)].font = self.font1
            ws[cell_id + str(line)].alignment = Alignment(horizontal='center', vertical='center')
            line = line + 1

        ws.cell(row=line,column=id).value = '总计'
        cell_id = self.alphabet[id]
        ws[cell_id + str(line)].border = self.border_style
        ws[cell_id + str(line)].font = self.font1
        ws[cell_id + str(line)].alignment = Alignment(horizontal='center', vertical='center')

class asset_reg_data_week(object):
    filepath = ""
    title1 = "资产支持计划"
    title2 = "组合产品"
    lst1 = ["序号","产品编码","机构全称","机构简称","产品全称","登记状态","产品规模（亿元）","报送日期","登记日期","中保登总时长（工作日）","基础资产类型","涉及的信托计划备案类型是否为主动管理类","是否分期","是否分级","发行场所","产品期限","产品托管人","原始权益人","原始债务人","资产服务人","内部增信","内部增信描述","外部增信","外部增信描述"]
    lst2 = ["序号","流程进度","产品管理人","产品全称","报送时间","登记日期","目标投资者是否包含自然人","是否向保险资金募集","登记编码","是否涉及代销","代销机构名称","登记方式"]
    alphabet = ['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    fill_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid") 
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border_style = Border(
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'),
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000')
        )
    font1 = Font(bold=True, size=12)
    wb = None
    ws1 = None
    ws2 = None

    def __init__(self,f):
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.filepath = f

    def save_sheet(self):
        self.wb.active = self.ws1
        self.wb.save(self.filepath)
        print(self.filepath + " 文件已生成！")
        
    def set_ws1_sheet(self):
        self.ws1.title = self.title1
        self.set_excel_sheet_row_text(1,self.ws1,self.lst1,1,False,self.font1)
        self.set_merge_cells_border_row(self.ws1,"A1","I1")

    def set_ws2_sheet(self):
        self.ws2 = self.wb.create_sheet(self.title2)
        self.set_excel_sheet_row_text(1,self.ws2,self.lst2,1,False,self.font1)

    def set_merge_cells_border_row(self,ws,bstart,bend):
        #设置合并后的单元格边框
        start = bstart[0]
        end = bend[0]
        row = str(int(bstart[1:]))
        i = 1
        bstart = False
        while i < len(self.alphabet):
            if start == self.alphabet[i]:
                bstart = True
            if bstart and end == self.alphabet[i]:
                bstart = False
                ws[self.alphabet[i] + row].border = self.border_style
            
            if bstart == True:
                ws[self.alphabet[i] + row].border = self.border_style

            i = i + 1

    def set_excel_sheet_row_text(self,line,ws,text_list,id,fill_color = True,font = None):
        fill_grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for item in text_list:
            ws.cell(row=line,column=id).value = item
            cell_id = self.alphabet[id]
            ws.column_dimensions[cell_id].width = 15
            if fill_color:
                ws[cell_id + str(line)].fill = fill_grey
            if font is not None:
                ws[cell_id + str(line)].font = font
            ws[cell_id + str(line)].border = self.border_style
            if len(str(item)) <= 6:
                ws[cell_id + str(line)].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[cell_id + str(line)].alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')

            id = id + 1
    
def savelog(strsen,display = True):
    if display:
        print(strsen)
    try:
        resultfile = open("debug.log",'a+')
        resultfile.write("[" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "]" + strsen + "\n")
        resultfile.close()
    except Exception as e:
        pass

def update_tup(tup1):
    if len(tup1) == 8:
        update_maps_line(tup1[0],tup1[1],tup1[2],tup1[3],tup1[4],tup1[5],tup1[6],tup1[7])
    elif len(tup1) == 6:
        update_maps_line(tup1[0],tup1[1],tup1[2],tup1[3],tup1[4],tup1[5],None,None)
    else:
        print("update_tup 该元组长度不支持!")
        sys.exit(1)

def merge_maps_data(map1,map2):
    resmap = {}
    for key in map1:
        #key是管理人名字
        # 0:固定收益类
        # 1:权益类
        # 2:商品及金融衍生品类
        # 9:混合类
        v1map = map1[key]
        v0 = 0
        v1 = 0
        v2 = 0
        v9 = 0
        if "0" in v1map and str_is_int(v1map["0"]):
            v0 = int(v1map["0"])
        if "1" in v1map and str_is_int(v1map["1"]):
            v1 = int(v1map["1"])
        if "2" in v1map and str_is_int(v1map["2"]):
            v2 = int(v1map["2"])
        if "9" in v1map and str_is_int(v1map["9"]):
            v9 = int(v1map["9"])

        resmap[key] = {"fixed":v0,"right":v1,"product_finance":v2,"mixed":v9,"open":0,"close":0}

    for key in map2:
        #key是管理人名字
        # 1:开放式
        # 2:封闭式
        v2map = map2[key]
        v1 = 0
        v2 = 0
        if "1" in v2map and str_is_int(v2map["1"]):
            v1 = int(v2map["1"])
        if "2" in v2map and str_is_int(v2map["2"]):
            v2 = int(v2map["2"])

        if key in resmap:
            resmap[key]["open"] = v1
            resmap[key]["close"] = v2
        else:
            resmap[key] = {"fixed":0,"right":0,"product_finance":0,"mixed":0,"open":v1,"close":v2}
    
    return resmap


def update_maps_line(map1,map2,map3,map4,map5,map6,map7,map8):
    #统一各个字典中对应项目的行的位置信息
    m1 = {}
    m1.update(map1)
    m1.update(map2)
    m1.update(map3)
    m1.update(map4)
    m1.update(map5)
    m1.update(map6)
    if map7 is not None:
        m1.update(map7)
    if map8 is not None:
        m1.update(map8)

    line = 0
    for k in m1:
        m1[k] = line
        line = line + 1
    
    for k in map1:
        map1[k][1] = m1[k]
    for k in map2:
        map2[k][1] = m1[k]
    for k in map3:
        map3[k][1] = m1[k]
    for k in map4:
        map4[k][1] = m1[k]
    for k in map5:
        map5[k][1] = m1[k]
    for k in map6:
        map6[k][1] = m1[k]
    
    if map7 is not None:
        for k in map7:
            map7[k][1] = m1[k]
    
    if map8 is not None:
        for k in map8:
            map8[k][1] = m1[k]

def get_row(map,key):
    maxrow = 0
    for k in map:
        r = map[k][1]
        if k == key:
            return r
        if int(r) > maxrow:
            maxrow = int(r)

    return maxrow

def get_max_row(map):
    maxrow = 0
    for k in map:
        r = map[k][1]
        if int(r) > maxrow:
            maxrow = int(r)
    return maxrow

def get_max_row_maps(map1,map2):
    max1 = get_max_row(map1)
    max2 = get_max_row(map2)
    return max1 + max2

def get_list_max_value(lst):
    lst.sort()
    return lst[-1] + 1

def str_is_int(v):
    try:
        int(v)
        return True
    except ValueError:
        return False
    
def str_is_float(v):
    try:
        float(v)
        return True
    except ValueError:
        return False

def get_format_value(v):
    if str(v) == "0.0" or str(v) == "0":
        return 0
    elif str_is_int(v):
        return int(v)
    elif str_is_float(v):
        return float(v)
    else:
        return v

def get_default_value(value,default_value=""):
    if value is None:
        return default_value
    else:
        return value

def get_truncate_date(datestr):
    if datestr is None or str(datestr) == "None":
        return ""
    datestr = datestr.strip()
    if len(str(datestr)) == 0:
        return ""
    if datestr.find(' ') == -1:
        return datestr
    else:
        return datestr.split(' ')[0]


def get_increase_str(pre_data,cur_data):
    #获得增长的百分比
    if pre_data is None or cur_data is None or len(str(cur_data)) == 0 or len(str(pre_data)) == 0 or float(pre_data) == 0.0:
        return ""
    else:
        return str(round(((float(cur_data) - float(pre_data)) / float(pre_data)) * 100)) + "%"

def switch_reg_date(input_time_str):
    # 将传入的时间字符串转换为 datetime 对象
    input_time = datetime.datetime.strptime(input_time_str, "%Y-%m-%d %H:%M:%S.%f")
    # input_time = input_time_str.strftime('%Y-%m-%d %H:%M:%S')
    # 判断输入的时间是否大于 15:00:00
    if input_time.hour >= 15:
        # 如果时间大于 15:00:00，则返回第二天的日期
        next_day = input_time + datetime.timedelta(days=1)
        return next_day.date()  # 返回第二天的日期，类型为日期
    else:
        # 如果时间小于 15:00:00，则返回当前日期
        return input_time.date()  # 返回当前日期，类型为日期

def list_dates_between(date_str1, date_str2):
    if len(date_str1) == 0 or len(date_str2) == 0:
        return [] 
    date_str1 = get_truncate_date(date_str1)
    date_str2 = get_truncate_date(date_str2)
    date1 = datetime.datetime.strptime(date_str1, "%Y-%m-%d")
    date2 = datetime.datetime.strptime(date_str2, "%Y-%m-%d")
    start_date = None
    end_date = None
    if date1 > date2:
        start_date = datetime.datetime.strptime(date_str2, "%Y-%m-%d")
        end_date = datetime.datetime.strptime(date_str1, "%Y-%m-%d")
    else:
        start_date = datetime.datetime.strptime(date_str1, "%Y-%m-%d")
        end_date = datetime.datetime.strptime(date_str2, "%Y-%m-%d")

    # 存放结果日期的列表
    date_list = []
    
    # 循环从start_date到end_date
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date.date())  # 只取日期部分
        current_date +=datetime.timedelta(days=1)  # 增加一天
    
    return date_list

def check_date(year, month, day):
    try:
        datetime.datetime(year, month, day).date()
    except ValueError:
        print("日期不合法:" + str(year) + "-" + str(month) + "-" + str(day))
        sys.exit(1)

def get_date_info_map(year = "",month = "",day = ""):
    year_date_start = 0
    year_date_end = 0

    quarter_date_start = 0
    quarter_date_end = 0

    month_date_start = 0
    month_date_end = 0
    
    pre_year_date_start = 0
    pre_year_date_end = 0

    pre_quarter_date_start = 0
    pre_quarter_date_end = 0

    pre_month_date_start = 0
    pre_month_date_end = 0

    cur_year = ""
    cur_month = ""
    cur_day = ""
    if len(year) == 0 or len(month) == 0 or len(day) == 0:
        cur_year = (date.today()).strftime('%Y')
        cur_month = (date.today()).strftime('%m')
        cur_day = (date.today()).strftime('%d')
    else:
        cur_year = year
        if len(month) < 2:
            cur_month = "0" + month
        else:
            cur_month = month

        if len(day) < 2:
            cur_day = "0" + day
        else:
            cur_day = day

        check_date(int(cur_year),int(cur_month),int(cur_day))

    quarter = get_quarter_from_month(cur_month)
    init_date = "1900-01-01"
    cur_date = cur_year + "-" + cur_month + "-" + cur_day

    pre_year_cur_date = str(int(cur_year) - 1) + '-'
    if cur_month == '02' and cur_day == '29':
        pre_year_cur_date = pre_year_cur_date + '02-28'
    else:
        pre_year_cur_date = pre_year_cur_date + cur_month + '-' + cur_day
    
    year_date_start = str(cur_year) + "-01-01"
    year_date_end = str(cur_year) + "-12-31"

    ql = get_quarter_date_from_month(cur_year,cur_month)
    quarter_date_start = ql[0]
    quarter_date_end = ql[1]

    ml = get_month_date_from_month(cur_year,cur_month)
    month_date_start = ml[0]
    month_date_end = ml[1]
    #############################################################
    pre_year_date_start = str(int(cur_year) - 1) + "-01-01"
    pre_year_date_end = str(int(cur_year) - 1) + "-12-31"

    pq = get_quarter_date_from_month(int(cur_year) - 1,cur_month)
    pre_year_quarter_date_start = pq[0]
    pre_year_quarter_date_end = pq[1]

    qp = get_pre_quarter_date_from_month(cur_year,cur_month)
    pre_quarter_date_start = qp[0]
    pre_quarter_date_end = qp[1]

    mp = get_pre_month_date_from_month(cur_year,cur_month)
    pre_month_date_start = mp[0]
    pre_month_date_end = mp[1]
    ############################################################
    pml = get_month_date_from_month(int(cur_year) - 1,cur_month)
    pre_year_month_start = pml[0]
    pre_year_month_end = pml[1]
    ############################################################
    #获得当前年的1-12月份的月份范围
    current_months_scope = []
    for i in range(13):
        if i == 0:
            continue
        month1 = get_month_date_from_month(int(cur_year),i)
        current_months_scope.append(month1)
    ############################################################
    savelog("截止到今天日期范围：" + str(init_date) + " " + str(cur_date))
    savelog("当前季度为：" + str(quarter))
    savelog("截至到去年今天的日期范围：" + str(init_date) + " " + str(pre_year_cur_date))
    savelog("今年日期范围：" + str(year_date_start) + " " + str(year_date_end))
    savelog("当前季度日期范围：" + str(quarter_date_start) + " " + str(quarter_date_end))
    savelog("当前月日期范围：" + str(month_date_start) + " " + str(month_date_end))
    savelog("前一个季度日期范围：" + str(pre_quarter_date_start) + " " + str(pre_quarter_date_end))
    savelog("前一个月日期范围：" + str(pre_month_date_start) + " " + str(pre_month_date_end))
    savelog("去年日期范围：" + str(pre_year_date_start) + " " + str(pre_year_date_end))
    savelog("去年当前月日期范围：" + str(pre_year_month_start) + " " + str(pre_year_month_end))
    savelog("去年当前季度日期范围：" + str(pre_year_quarter_date_start) + " " + str(pre_year_quarter_date_end))
    savelog("今年1-12月日期范围：" + str(current_months_scope))
    resmap = {
                "year":str(cur_year),"month":str(cur_month),"day":str(cur_day),"quarter":str(quarter),\
                "init_date":str(init_date),"cur_date":str(cur_date),"pre_year_cur_date":str(pre_year_cur_date),\
                "year_date_start":str(year_date_start),"year_date_end":str(year_date_end),\
                "quarter_date_start":str(quarter_date_start),"quarter_date_end":str(quarter_date_end),\
                "month_date_start":str(month_date_start),"month_date_end":str(month_date_end),\
                "pre_year_date_start":str(pre_year_date_start),"pre_year_date_end":str(pre_year_date_end),\
                "pre_quarter_date_start":str(pre_quarter_date_start),"pre_quarter_date_end":str(pre_quarter_date_end),\
                "pre_month_date_start":str(pre_month_date_start),"pre_month_date_end":str(pre_month_date_end),\
                "pre_year_month_start":str(pre_year_month_start),"pre_year_month_end":str(pre_year_month_end),\
                "pre_year_quarter_date_start":str(pre_year_quarter_date_start),"pre_year_quarter_date_end":str(pre_year_quarter_date_end),\
                "current_year_months_scope":current_months_scope
                }
    return resmap

def valid_month_quarter(quarter,month):
    #判断月份和季度是否匹配
    if (int(month) >= 1 and int(month) <= 3 and int(quarter) == 1) or \
        (int(month) >= 4 and int(month) <= 6 and int(quarter) == 2) or \
        (int(month) >= 7 and int(month) <= 9 and int(quarter) == 3) or \
        (int(month) >= 10 and int(month) <= 12 and int(quarter) == 4):
        return True
    else:
        return False

def get_quarter_from_month(month):
    #根据月份获得当前季度
    if int(month) >= 1 and int(month) <= 3:
        return 1
    elif int(month) >= 4 and int(month) <= 6:
        return 2
    elif int(month) >= 7 and int(month) <= 9:
        return 3
    elif int(month) >= 10 and int(month) <= 12:
        return 4
    else:
       return -1

def get_month_date_from_quarter(year,quarter):
    #根据季度获得当前开始结束日期
    start = ""
    end = ""
    if int(quarter) == 1:
        start = "-01-01"
        end = "-03-31"
    elif int(quarter) == 2:
        start = "-04-01"
        end = "-06-30"
    elif int(quarter) == 3:
        start = "-07-01"
        end = "-09-30"
    else:
        start = "-10-01"
        end = "-12-31"
    
    return [str(year) + start,str(year) + end]

def get_pre_month_date_from_quarter(year,quarter):
    #根据当前季度获得前一季度的开始结束日期
    if int(quarter) == 1:
        return get_month_date_from_quarter(int(year) - 1,4)
    else:
        return get_month_date_from_quarter(year,int(quarter) - 1)

def get_quarter_date_from_month(year,month):
    #根据月份获得当前季度的开始结束日期
    start = ""
    end = ""
    if int(month) >= 1 and int(month) <= 3:
        start = "-01-01"
        end = "-03-31"
    elif int(month) >= 4 and int(month) <= 6:
        start = "-04-01"
        end = "-06-30"
    elif int(month) >= 7 and int(month) <= 9:
        start = "-07-01"
        end = "-09-30"
    else:
        start = "-10-01"
        end = "-12-31"
    
    return [str(year) + start,str(year) + end]

def get_pre_quarter_date_from_month(year,month):
    #根据当前月份获得前一个季度的开始结束日期
    if int(month) <= 3:
        #当前是月份小于等于3，属于第一季度，返回上一年10月份所在季度(第四季度)的开始结束日期
        return get_quarter_date_from_month(int(year) - 1,10)
    elif int(month) <= 6:
        #当前是月份大于3小于等于6，属于第二季度，返回当年1月份所在季度(第一季度)的开始结束日期
        return get_quarter_date_from_month(year,1)
    elif int(month) <= 9:
        #当前是月份大于6小于等于9，属于第三季度，返回当年4月份所在季度(第二季度)的开始结束日期
        return get_quarter_date_from_month(year,4)
    else:
        #当前是月份大于9小于等于12，属于第四季度，返回当年7月份所在季度(第三季度)的开始结束日期
        return get_quarter_date_from_month(year,7)
    
def get_month_date_from_month(year,month):
    #根据年份 获得当前月份的开始结束日期
    if len(str(month)) == 1:
        month = "0" + str(month)

    date_start = str(year) + "-" + str(month) + "-01"
    date_end = ""
    if int(month) == 1 or int(month) == 3 or int(month) == 5 or int(month) == 7 or int(month) == 8 or int(month) == 10 or int(month) == 12:
        date_end = str(year) + "-" + str(month) + "-31"
    elif int(month) == 4 or int(month) == 6 or int(month) == 9 or int(month) == 11:
        date_end = str(year) + "-" + str(month) + "-30"
    else:
        if calendar.isleap(int(year)) == True:
            date_end = str(year) + "-" + str(month) + "-29"
        else:
            date_end = str(year) + "-" + str(month) + "-28"

    return [date_start,date_end]

def get_pre_month_date_from_month(year,month):
    #根据当前月份获得前一个月的开始结束日期
    if int(month) == 1:
        return get_month_date_from_month(int(year) - 1,"12")
    else:
        return get_month_date_from_month(year,str(int(month) - 1))


def read_config():
    #读取配置文件
    configfile = "config.ini"
    resmap = {}
    resmap["attachment1"] = "a1.xlsx"
    resmap["year"] = ""
    resmap["month"] = ""
    resmap["day"] = ""
    resmap["oracle_base"] = "E:\\app\\hspcadmin\\product"
    resmap["oracle_sid"] = "kfzgdb"
    resmap["oracle_home"] = resmap["oracle_base"] + '\\11.2.0\\client_1'
    resmap["ld_library_path"] = resmap["oracle_base"] + '\\11.2.0\\client_1\\LIB'
    localfile = None
    try:
        localfile = open(configfile,"r",encoding='utf-8')
        allline   = localfile.readlines()
        for line in allline:
            if len(line) == 0 or line[0] == '#':
                continue
            if line.find('=') != -1:
                line = line.replace("\"","").replace('\n','').replace('\r','')
                s1 = line.split('=')
                if s1[0].strip() == "attachment1":
                    resmap["attachment1"] = s1[1].strip()
                    continue

                if s1[0].strip() == "date1":
                    config_date = s1[1].strip()
                    if config_date.find('-') != -1:
                        da = config_date.split('-')
                        resmap["year"] = str(da[0])
                        resmap["month"] = str(da[1])
                        resmap["day"] = str(da[2])
                        check_date(int(da[0]),int(da[1]),int(da[2]))
                        continue

                if s1[0].strip() == "oracle_base":
                    if len(s1[1]) == 0:
                        print("需要配置oracle_base目录,例如E:\\app\\hspcadmin\\product")
                        sys.exit(1)
                    else:
                        resmap["oracle_base"] = s1[1].strip()
                    continue

                if s1[0].strip() == "oracle_sid":
                    if len(s1[1]) == 0:
                        print("需要配置oracle_sid,例如kfzgdb")
                        sys.exit(1)
                    else:
                        resmap["oracle_sid"] = s1[1].strip()
                    continue

                if s1[0].strip() == "oracle_home":
                    if len(s1[1]) == 0:
                        print("需要配置oracle_home")
                        sys.exit(1)
                    else:
                        resmap["oracle_home"] = s1[1].strip()
                    continue

                if s1[0].strip() == "ld_library_path":
                    if len(s1[1]) == 0:
                        print("需要配置ld_library_path")
                        sys.exit(1)
                    else:
                        resmap["ld_library_path"] = s1[1].strip()
                    continue

    except OSError as reason:
        savelog("配置文件不存在，使用默认参数!")
    finally:
        if localfile is not None:
            localfile.close()

    savelog("读取配置文件:"+str(resmap))
    return resmap
