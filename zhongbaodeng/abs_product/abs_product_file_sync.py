# coding=utf-8
# !/usr/bin/python
#-------------------------------------------------------------------------------
# Name:        abs_product_file_sync
# Purpose:     同步资产资产计划abs提交受理处理后的附件信息
#
# Author:      lihuaizhong
#
# Created:     10-11-2020
# Copyright:   (c) lihuaizhong 2020
# Licence:     <your licence>
#-------------------------------------------------------------------------------
# 获取估值需要的产品信息列表和附件

import logging
import os
import os.path
import socket
import sys
import time

import paramiko
from openpyxl import Workbook
from docxtpl import DocxTemplate
try:
    import cx_Oracle
except ImportError as e:
    print("导入cx_Oracle模块失败,需要安装cx_Oracle\r\n{}",str(e))
    sys.exit(-1)

logging.basicConfig()

# 附件catalog
catalog = 'TFUNDINFO'

class SSHSession(object):

    def __init__(self,fileSrcHost,fileSrcPort,fileSrcUserName,fileSrcPassword,dbip,dbport,dbSID,dbUserName,dbPassword,key_file=None):

        self.fileSrcHost = fileSrcHost
        self.fileSrcPort = fileSrcPort
        self.fileSrcUserName = fileSrcUserName
        self.fileSrcPassword = fileSrcPassword

        self.dbip = dbip
        self.dbport = dbport
        self.dbSID = dbSID
        self.dbusername = dbUserName
        self.dbpassword = dbPassword

        self.fileSrcSftp = self.getFtp(self.fileSrcHost,self.fileSrcPort,self.fileSrcUserName,self.fileSrcPassword)
        self.cursor = self.getDbCursor(self.dbip,self.dbport,self.dbSID,self.dbusername,self.dbpassword)

    def getDbCursor(self,dbIp,dbPort,dbSID,dbUserName,dbPassword):
        try:
            dsn_tns = cx_Oracle.makedsn(dbIp, dbPort, service_name=dbSID)
            self.db = cx_Oracle.connect(dbUserName, dbPassword, dsn_tns)
            cursor = self.db.cursor()
            print("connet %s:%s/%s database successful!" % (dbIp,dbPort,dbSID))
            return cursor
        except cx_Oracle.Error as e:
            print("init connet database failed,exit!" + str(e))
            time.sleep(5)
            sys.exit(-1)

    def getFtp ( self ,host,port,userName,userPassword):
        print("connect fileserver %s:%s" % (host,port))
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(30 * 60 * 1000)
        sock.connect((host, port))
        t = paramiko.Transport(sock)
        t.start_client()
        # supposed to check for key in keys, but I don't much care right now to find the right notation
        if userPassword is not None:
            t.auth_password(userName, userPassword, fallback=False)
        else:
            raise Exception('Must supply either key_file or password')
        print("connect fileserver %s:%s Successful!!" % (host, port))
        return paramiko.SFTPClient.from_transport(t)

    def reconnectDB(self):
        try:
            print("reconnet %s:%s/%s database " % (self.dbip,self.dbport,self.dbSID))
            self.cursor = self.getDbCursor(self.dbip,self.dbport,self.dbSID,self.dbUserName,self.dbPassword)
            print("reconnet %s:%s/%s database successful!" % (self.dbip,self.dbport,self.dbSID))
        except cx_Oracle.Error as e:
            print("reconnet database failed,exit!" + str(e))
            time.sleep(5)
            sys.exit(-1)

    def disconnect(self):
        try:
            self.fileSrcSftp.close()
            self.db.close()
            print("Connection terminated.")
        except cx_Oracle.DatabaseError as e:
            print("close db  except." + str(e))

    # 获取产品代码：为产品注册流程的分配受理节点提交受理通过后的产品代码
    def getFundcodes(self,templetId,startDate,endDate,localpath):
        fundcodes = set()
        sql = "select distinct f.c_fundcode,f.c_fundname from tfundinfo f where f.l_pdttempletid ='"+templetId+"' and f.c_fundcode in (select j.BIZ_FLOW_UUID_  from jbpm4_ext_htask_ j  where j.PROC_NAME_='产品注册流程'  and j.ACTIVITY_NAME_='分配受理' and j.TASKOUTCOMEACTION='提交受理' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>='"+startDate+"' and to_char(j.END_,'yyyymmdd')<='"+endDate+"' \
               and  j.BIZ_FLOW_UUID_ not in (select j.BIZ_FLOW_UUID_  from jbpm4_ext_htask_ j where j.PROC_NAME_='产品注册流程'  and j.ACTIVITY_NAME_='中保登受理' and to_char(j.END_,'yyyymmdd')>='"+startDate+"' and to_char(j.END_,'yyyymmdd')<='"+endDate+"')) "
        print('getFundcode.startDate = '+ startDate+',endDate='+endDate+',templetName='+ self.getTempleName(templetId),'sql='+sql)
        try:
            dataResult = self.cursor.execute(sql)
            data = dataResult.fetchall()
            for r in range(len(data)):
                fundcode = data[r][0]
                tafundcode = data[r][1]
                fundcodes.add(fundcode+','+tafundcode)
            return fundcodes

        except Exception as e:
            print('获取templetId=' + templetId + '类型的产品信息异常'+str(e))
        return fundcodes
    #获取组合类统计报告
    def getCombProductReport(self,startDate,endDate,fileLocalTemplateFile,localpath):
        sql = "select substr("+endDate+",1,4) reportYear,substr("+endDate+",5,2) reportMonth,to_number(substr("+endDate+",7,2)) reportDay,\
         a.mgrorgnoCount,a.combproductCount,a.sucessCount,b.per24hCount,c.nonAcceptCount,a.backCount,d.setreportCount,\
         a.fixedincomeCount,a.fixedIncomeRatio,a.mixCount,a.rightsCount, \
         a.mixRatio,a.rightsRatio,a.structuredCount,a.nonStructuredCount,e.istoinsureassetCount, \
         e.isincludepersonalCount,e.isfirstCount,e.ismonCount,e.ismoneyflagCount,F.isAutoCacelCount,F.isApplyCacelCount,F.cacelCount,e.isfofCount from ( \
   /*产品管理人,报送产品总量*/  \
      select count(distinct b.c_mgrorgno) mgrorgnoCount,count(distinct f.c_fundcode) combproductCount,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',f.c_fundcode)) sucessCount,\
          count(distinct decode(j.TASKOUTCOMEACTION,'不予登记',f.c_fundcode)) backCount,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'0',f.c_fundcode))) fixedincomeCount,\
          round(decode(count(distinct f.c_fundcode),0,0,count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'0',f.c_fundcode)))/count(distinct decode(j.TASKOUTCOMEACTION, '准予登记', f.c_fundcode)))*100, 2)||'%' fixedIncomeRatio,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'9',f.c_fundcode ))) mixCount,\
          round(decode(count(distinct f.c_fundcode),0,0,count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'9',f.c_fundcode )))/count(distinct decode(j.TASKOUTCOMEACTION, '准予登记', f.c_fundcode)))*100, 2)||'%' mixRatio,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'1',f.c_fundcode ))) rightsCount,\
          round(decode(count(distinct f.c_fundcode),0,0,count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(C_PRDTYPE,'1',f.c_fundcode )))/count(distinct decode(j.TASKOUTCOMEACTION, '准予登记', f.c_fundcode)))*100, 2)||'%' rightsRatio,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(f.l_pdttempletid,'10',f.c_fundcode))) nonStructuredCount,\
          count(distinct decode(j.TASKOUTCOMEACTION,'准予登记',decode(f.l_pdttempletid,'17',f.c_fundcode))) structuredCount\
    from tfundinfo f ,tpdtbasicinfo  b,(select j.BIZ_FLOW_UUID_ c_fundcode,j.TASKOUTCOMEACTION  from jbpm4_ext_htask_ j  where j.PROC_NAME_='发行前登记流程' and j.ACTIVITY_NAME_ in ( '中保登审批','中保登复核')\
   and j.OUTCOME_ in ('pass','oppose') and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+") j where f.l_pdttempletid in ('10','17') \
   and f.c_fundcode=b.c_fundcode and  f.c_fundcode=j.c_fundcode)a,(\
   /*24小时内予以登记的产品*/ \
   select count(distinct a. BIZ_FLOW_UUID_) per24hCount from (\
   select j.BIZ_FLOW_UUID_,max(j.END_) END_  from jbpm4_ext_htask_ j  where j.PROC_NAME_='发行前登记流程' and j.ACTIVITY_NAME_ in ( '中保登审批','中保登复核')\
   and j.TASKOUTCOMEACTION='准予登记' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+" group by j.BIZ_FLOW_UUID_)a,(\
    select j.BIZ_FLOW_UUID_,min(j.CREATE_ )CREATE_ from jbpm4_ext_htask_ j  where j.PROC_NAME_='发行前登记流程' and j.ACTIVITY_NAME_='分配受理'\
   and j.TASKOUTCOMEACTION='提交受理' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+"  group by j.BIZ_FLOW_UUID_  )b\
   where a.BIZ_FLOW_UUID_=b.BIZ_FLOW_UUID_ and TO_NUMBER((cast(a.END_ AS DATE)- cast(b.CREATE_ AS DATE))*24)<=24)b,(\
   /*不予受理的产品*/ \
   select count(distinct BIZ_FLOW_UUID_) nonAcceptCount  from jbpm4_ext_htask_ j  where j.PROC_NAME_='发行前登记流程' and j.ACTIVITY_NAME_='不予受理复核'\
   and j.TASKOUTCOMEACTION='复核通过' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+" )c,(\
   /*产品设立报告*/ \
   select count(distinct f.c_fundcode)  setreportCount from  tfundinfo f ,tblobstorage bs where f.l_pdttempletid in ('10','17') and  f.c_fundcode in (\
   select j.BIZ_FLOW_UUID_  from jbpm4_ext_htask_ j  where j.PROC_NAME_ in ( '产品设立报告流程','设立报告提交流程') and j.ACTIVITY_NAME_ in  ('产品设立报告审核', '平台接收')\
   and j.TASKOUTCOMEACTION='接收' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+" )\
   and f.c_fundcode=bs.c_subkeyid and bs.c_memo='产品设立报告')d,(\
   /*MOM产品*/ \
    select\
           count(decode(b.C_ISTOINSUREASSET,'0',f.c_fundcode,null  )) istoinsureassetCount ,/*面向非保险资金募集只数*/ \
           count(decode(b.C_ISINCLUDEPERSONAL,'1',f.c_fundcode,null  )) isincludepersonalCount,/*面向自然人投资者募集的产品只数*/ \
           count(decode(b.C_ISFIRSTPRODUCT,'1',f.c_fundcode,null  )) isfirstCount,/*首单产品只数*/ \
           count(decode(b.C_ISMMTYPE,'1',f.c_fundcode,null  )) ismonCount,/*MOM产品只数*/ \
           count(decode(b.C_ISFFTYPE,'1',f.c_fundcode,null  )) isfofCount,/*FOF产品只数*/ \
           count(decode(b.C_ISMONEYFLAG,'1',f.c_fundcode,null )) ismoneyflagCount/*货币市场类产品只数*/ \
    from tfundinfo f ,TPDTASSORTEDPRODUCTINFO  b where f.l_pdttempletid in ('10','17') and  f.c_fundcode in (\
   select j.BIZ_FLOW_UUID_  from jbpm4_ext_htask_ j  where j.PROC_NAME_='发行前登记流程' and j.ACTIVITY_NAME_ in ( '中保登审批','中保登复核')\
   and j.TASKOUTCOMEACTION='准予登记' and j.OUTCOME_='pass' and to_char(j.END_,'yyyymmdd')>="+startDate+" and to_char(j.END_,'yyyymmdd')<="+endDate+" )\
   and f.c_fundcode=b.c_fundcode)e,( SELECT \
          COUNT(DECODE(C.C_CANCELTYPE,'0',A.C_FUNDCODE)) isAutoCacelCount, \
          COUNT(DECODE(C.C_CANCELTYPE,'1',A.C_FUNDCODE)) isApplyCacelCount, \
          COUNT(A.C_FUNDCODE) cacelCount \
   FROM TFUNDINFO A JOIN TPDTBASICINFO B ON A.C_FUNDCODE = B.C_FUNDCODE JOIN TPDTEXTENDINFO C ON A.C_FUNDCODE = C.C_FUNDCODE \
   WHERE  A.C_CLIENTAUDITFLAG = 48 AND TO_CHAR(B.D_CANCELDATE,'yyyymmdd') <= "+endDate+") F"
        print('getCombProductReport sql='+sql)
        try:
            if not os.path.isfile(fileLocalTemplateFile):
                print('模板文件不存在:'+fileLocalTemplateFile)
                exit()
            tpl = DocxTemplate(fileLocalTemplateFile)
            dataResult = self.cursor.execute(sql)
            data = dataResult.fetchall()
            for r in range(len(data)):

                context = {'year': data[r][0],'month': data[r][1],'day': data[r][2],'mgrorgnoCount': data[r][3],
                           'combproductCount': data[r][4],'sucessCount': data[r][5],'per24hCount': data[r][6],'nonAcceptCount': data[r][7],
                           'backCount': data[r][8],'setreportCount': data[r][9],'fixedincomeCount': data[r][10],'fixedIncomeRatio': data[r][11],
                           'mixCount': data[r][12],'rightsCount': data[r][13],'mixRatio': data[r][14],'rightsRatio': data[r][15],
                           'structuredCount': data[r][16],'nonStructuredCount': data[r][17],'istoinsureassetCount': data[r][18],'isincludepersonalCount': data[r][19],
                           'isfirstCount': data[r][20],'ismonCount': data[r][21],'ismoneyflagCount': data[r][22],'isAutoCacelCount': data[r][23],'isApplyCacelCount': data[r][24],'cacelCount': data[r][25],
                           'isfofCount': data[r][26]
                          }
            tpl.render(context)
            #tmpreportpath = localpath +os.path.sep +'combreport'+ os.path.sep
            tmpreportpath = localpath +os.path.sep + endDate + os.path.sep + 'annex'+ os.path.sep
            self.mkdir_ifnotexists(tmpreportpath)
            tpl.save(tmpreportpath+'Comb_product_Report_'+endDate+'.docx')
            print('生成'+tmpreportpath+' 成功')
        except Exception as e:
            print('获取组合类型的产品信息异常'+str(e))
    # 获取融资主体/担保主体编码
    def getMainCode(self,name):
        if name is None:
            return ''
        if '' == name.strip():
            return name.strip()
        # sql = "  select C_ORGCODE from TPDTMANDATOR where c_name = '"+name+"'"
        sql = " select C_ORGCODE from TPDTMANDATOR where c_status not in ('0','2') and c_name ='"+name+"' order by length(c_orgcode) desc"
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        if len(data) > 0:
            return data[0][0]
        else:
            return ''

    # 保存数据到excel(不包含表头)
    def write_to_excel(self,path,date,data,templetId):
        print('保存产品"'+self.getTempleName(templetId)+'"详细信息到excel开始！')
        self.mkdir_ifnotexists(path+os.path.sep+date)
        os.chdir(path+os.path.sep+date)

        file_name = '{0}_{1}.xlsx'.format(date,self.getTempleName(templetId))
        wb = Workbook()
        sheet = wb.active
        sheet.title=date

        for r in range(len(data)):
            for c in range(len(data[r])):
                sheet.cell(row = r+1,column = c+1).value=data[r][c]
        wb.save(file_name)
        print('保存产品"'+self.getTempleName(templetId)+'"详细信息到excel结束！')

    # 保存数据到excel同时保存表头
    def write_to_excel_title(self,path,date,data,titles,templetId):
        print('保存产品详细信息到excel开始！')
        if len(data) > 0:
            self.mkdir_ifnotexists(path+os.path.sep+date)
            os.chdir(path+os.path.sep+date)
            file_name = '{0}_{1}.xlsx'.format(date,templetId)
            wb = Workbook()
            sheet = wb.active
            sheet.title=date

            for r in range(len(titles)):
                sheet.cell(row=1, column=r + 1).value = titles[r][0]

            for r in range(len(data)):
                for c in range(len(data[r])):
                    sheet.cell(row = r+2,column = c+1).value=data[r][c]
            wb.save(file_name)
        else:
            print('当天无‘'+templetId+'’产品数据！')
        print('保存产品详细信息到excel结束！')

    # 获取所有附件信息
    def read_tblobstorage_table(self,fileremotepath,filelocalpath,fundcode,startDate,tafundcode):
        print('【获取产品'+fundcode+'_'+tafundcode+'附件信息开始】')
        named_params = { 'subkeyid':fundcode}
        # sql = "SELECT  a.c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename FROM tblobstorage bs,tfundinfo a  WHERE bs.c_subkeyid=a.c_fundcode and c_catalog=:ccatalog AND c_subkeyid=:subkeyid"
        sql = " SELECT  a.c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename " \
              " FROM tblobstorage bs,tfundinfo a  " \
              " WHERE bs.c_subkeyid=a.c_fundcode  and bs.c_catalog='TFUNDINFOT' " \
              " AND c_subkeyid=:subkeyid "
        dataresult = self.cursor.execute(sql, named_params)
        data = dataresult.fetchall()
        for r in range(len(data)):
            print("\ttafundcode=%s, prdName=%s ,remotePath=%s ,localFileName=%s"%(tafundcode,data[r][0],data[r][1],data[r][2]) )
            self.getFileFromRemoteServer(filelocalpath,fileremotepath,tafundcode,data[r][0],data[r][1],data[r][2],startDate)
        print('【产品'+fundcode+'_'+tafundcode+'共获取'+str(len(data))+'条附件记录】')

    # 从远程文件服务器获取附件信息
    def getFileFromRemoteServer(self,localpath,remotepath,tafundcode,prdname,refilename,lcfilename,dataDate):
        tmpprdpath = localpath +os.path.sep + dataDate + os.path.sep + 'annex'+ os.path.sep + tafundcode
        self.mkdir_ifnotexists(tmpprdpath)
        localtmpFile = tmpprdpath+os.path.sep+lcfilename
        remotefile=remotepath+'/'+refilename
        print('\t\tremotefile='+remotefile+'\t,localtmpFile='+localtmpFile)
        try:
            self.fileSrcSftp.get(remotefile, localtmpFile)
        # shutil.copy(localfile,localtmpFile)
        except Exception as e:
            print('\t\t【获取附件异常,】' + str(e))

    # 获取存续期兑付附件信息
    def read_tblobstorage_table_Cash(self,fileremotepath,filelocalpath,lid,startDate,tafundcode):
        print('【获取产品'+tafundcode+'附件信息开始】')
        named_params = { 'csubkeyid':lid}
        # sql = "SELECT  a.c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename FROM tblobstorage bs,tfundinfo a  WHERE bs.c_subkeyid=a.c_fundcode and c_catalog=:ccatalog AND c_subkeyid=:subkeyid"
        sql = " SELECT  'nan' c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename " \
              " FROM tblobstorage bs   " \
              " WHERE (bs.c_catalog='TPDTDURATIONPROGRAM' )  " \
              " and c_subkeyid=:csubkeyid "
        print('sql2='+sql)
        dataresult = self.cursor.execute(sql, named_params)
        data = dataresult.fetchall()
        for r in range(len(data)):
            print("\ttafundcode=%s, prdName=%s ,remotePath=%s ,localFileName=%s"%(tafundcode,data[r][0],data[r][1],data[r][2]) )
            self.getFileFromRemoteServer_Cash(filelocalpath,fileremotepath,tafundcode,data[r][0],data[r][1],data[r][2],startDate)
        print('【产品'+tafundcode+'共获取'+str(len(data))+'条附件记录】')

    # 从远程文件服务器获取附件信息
    def getFileFromRemoteServer_Cash(self,localpath,remotepath,tafundcode,prdname,refilename,lcfilename,dataDate):
        tmpprdpath = localpath +os.path.sep + dataDate + os.path.sep + 'annex-cash'+ os.path.sep + tafundcode
        self.mkdir_ifnotexists(tmpprdpath)
        localtmpFile = tmpprdpath+os.path.sep+lcfilename
        remotefile=remotepath+'/'+refilename
        print('\t\tremotefile='+remotefile+'\t,localtmpFile='+localtmpFile)
        try:
            self.fileSrcSftp.get(remotefile, localtmpFile)
        # shutil.copy(localfile,localtmpFile)
        except Exception as e:
            print('\t\t【获取附件异常,】' + str(e))

    # 创建目录
    def mkdir_ifnotexists(self, path):
        path=path.strip()
        path=path.rstrip("\\")
        isExists=os.path.exists(path)

        # 判断结果
        if not isExists:
            # 如果不存在则创建目录
            os.makedirs(path)

            print('\t'+path+' 创建成功')
            return True
        else:
            # 如果目录存在则不创建，并提示目录已存在
            # print(path+' 目录已存在')
            return False

    # 从产品中心获取存续期产品
    def getDurationFundinfo(self,startDate,endDate,filelocalpath):
        print('从产品中心获取存续期startDate='+startDate+'，endDate='+endDate+'数据开始')
        sql="select  t.l_id 序号,  a.c_fundcode   产品代码, a.c_tafundcode TA产品代码, a.c_fundname 产品名称, decode( a.L_PDTTEMPLETID,'6','债权投资计划','8','资产支持计划') 产品类型," \
            " decode( t.c_type,'0','收益分配','1','赎回','2','回售','3','分期还本','4','兑付','5','清算','6','分红') 业务类型,  c_date 业务发生日期 ,  f_totalamount 业务发生总额, s_severaltime 期次," \
            " decode( c_paymenttype,'1','正常到期','2','提前还款')  兑付类型,  f_totalshare 业务发生总份数, t.d_createdate 创建日期,  f_afterfacevalue 执行后份额面值,  t.c_memo 备注" \
            " FROM tpdtdurationprogram t, tfundinfo a, member.operators o" \
            " WHERE t.c_fundcode=a.c_fundcode and  (a.L_PDTTEMPLETID in('6','8'))" \
            "   and t.c_creator = o.oper_id(+)" \
            "   and t.c_type='4'" \
            " and to_char(t.l_id) in (select biz_flow_uuid_ from jbpm4_ext_htask_ j where j.PROC_NAME_ like  '%存续期方案流程%'"\
            " and j.TASKOUTCOMEACTION = '复审通过' and j.activity_name_ = '存续期方案复审' "\
            " and to_char(j.END_, 'yyyyMMdd') >= '"+startDate+"' and to_char(j.END_, 'yyyyMMdd') <= '"+endDate+"' )"\
            " ORDER BY  t.c_type asc, t.d_createdate desc"
        dataResult = self.cursor.execute(sql)
        # 将产品信息存放到excel里
        dataRecords = dataResult.fetchall()
        self.write_to_excel_title(filelocalpath,endDate,dataRecords,dataResult.description,'存续期-兑付')
        print('从产品中心获取存续期startDate=' + startDate + '，endDate=' + endDate + '数据结束')

    # 从产品中心获取当日存续期兑付产品编码
    def getDurationFundinfo_Cash(self,templetId,startDate,endDate,filelocalpath):
        print('从产品中心获取存续期startDate='+startDate+'，endDate='+endDate+'数据开始')
        sql="select  to_char(t.l_id) as l_id,a.c_tafundcode "\
            " FROM tpdtdurationprogram t, tfundinfo a, member.operators o" \
            " WHERE t.c_fundcode=a.c_fundcode and a.L_PDTTEMPLETID= '"+templetId+"' " \
            "   and t.c_creator = o.oper_id(+)" \
            "   and t.c_type='4' " \
            " and to_char(t.l_id) in (select biz_flow_uuid_ from jbpm4_ext_htask_ j where j.PROC_NAME_ like  '%存续期方案流程%'"\
            " and j.TASKOUTCOMEACTION = '复审通过' and j.activity_name_ = '存续期方案复审' "\
            " and to_char(j.END_, 'yyyyMMdd') >= '"+startDate+"' and to_char(j.END_, 'yyyyMMdd') <= '"+endDate+"' )"\
            " ORDER BY  t.c_type asc, t.d_createdate desc"
        fundcodes=set()
        print('sql=' + sql)
        try:
            dataResult = self.cursor.execute(sql)
            # 将产品信息存放到excel里
            dataRecords = dataResult.fetchall()
            for r in range(len(dataRecords)):
                lid = dataRecords[r][0]
                ctafundcode = dataRecords[r][1]
                fundcodes.add(lid+','+ctafundcode)
            return fundcodes
            print('从产品中心获取存续期兑付startDate=' + startDate + '，endDate=' + endDate + '数据结束')
        except Exception as e:
            print('获取templetId=' + templetId + '类型的产品信息兑付异常'+str(e))
        return fundcodes

    # 从ta获取存续期产品信息
    def getFundInfoByTA(self,filelocalpath,startDate,endDate):
        print('从TA获取存续期startDate=' + startDate + '，endDate=' + endDate + '数据开始')
        sql = "SELECT a.c_fundcode 产品代码," \
                      " b.c_bondfullname 产品名称," \
                      " decode(b.c_producttype, '1', '债权', '2', '股权', '3', '资产支持计划') 产品类型," \
                      " decode(c_rdmflag, 'l', '回售', 'm', '赎回', 'i', '分期还本') 业务类型," \
                      " a.d_rdmdate 业务发生日期," \
                      " a.d_impdate 导入日期," \
                      " b.f_issueprice 执行后份额面值," \
                      " a.sumbalance 业务发生总额," \
                      " a.sumshare 业务发生总份数" \
            " FROM (SELECT a.c_fundcode, a.c_rdmflag, a.d_rdmdate, a.d_impdate," \
                      "    sum(a.f_redeembalance) sumbalance, sum(a.f_redeemshares) sumshare" \
                    " FROM tbatchrightline a" \
                    " group by a.c_fundcode, a.c_rdmflag, a.d_rdmdate, a.d_impdate) a, tbondinfo b" \
            " where a.c_fundcode = b.c_bondcode " \
              " and a.d_impdate >= '"+startDate+"' and a.d_impdate <= '"+endDate+"'" \
            " union all" \
            " SELECT a.c_fundcode 产品代码," \
                    " b.c_bondfullname 产品名称," \
                    " decode(b.c_producttype, '1', '债权', '2', '股权', '3', '资产支持计划') 产品类型," \
                    " '收益分配' 业务类型," \
                    " a.d_regdate 业务发生日期," \
                    " a.d_cdate 导入日期," \
                    " b.f_issueprice 执行后份额面值," \
                    " a.sumbalance 业务发生总额," \
                    " a.sumshare 业务发生总份数" \
            " FROM (SELECT a.c_fundcode, a.d_regdate, a.d_cdate," \
                    " sum(a.f_realbalance) sumbalance, sum(a.f_realshares) sumshare" \
                    " FROM tdividenddetail a" \
                    " group by a.c_fundcode, a.d_regdate, a.d_cdate) a,tbondinfo b" \
            " where a.c_fundcode = b.c_bondcode " \
                    " and a.d_cdate >= '"+startDate+"' and a.d_cdate <= '"+endDate+"'" \

        dataResult = self.cursor.execute(sql)
        # 将产品信息存放到excel里
        dataRecords = dataResult.fetchall()
        self.write_to_excel_title(filelocalpath, endDate, dataRecords, dataResult.description, '存续期-无兑付')
        print('从TA获取存续期startDate=' + startDate + '，endDate=' + endDate + '数据结束')

    def getTempleName(self,templetId):
        if templetId == '8':
            return '资产支持计划'
        elif templetId == '6':
            return '债权'
        elif templetId == '7':
            return '股权'
        else:
            return templetId

    def sendToDestFtp(self,localDir,date,fileDestHost,fileDestPort,fileDestUsername,fileDestPassword,fileDestRemotePath):
        fileDestSftp = self.getFtp(fileDestHost,fileDestPort,fileDestUsername,fileDestPassword)

        fileDestSftp.stat(fileDestRemotePath)
        stdin, stdout, stderr = fileDestSftp.exec_command('ls DIR')
        if stdout.readline() != '':
            print("exist")
        else:
            print("not exist")

        fileDestSftp.chdir(fileDestRemotePath)
        fileDestSftp.mkdir(date);
        fileDestSftp.put()

    def get_filelist(self,dir, Filelist):
        newDir = dir
        if os.path.isfile(dir):
            Filelist.append(dir)
        elif os.path.isdir(dir):
            for s in os.listdir(dir):
                newDir = os.path.join(dir, s)
                self.get_filelist(newDir, Filelist)
        return Filelist

    # 删除
    def del_file(self,path):
        if os.path.exists(path):
            ls = os.listdir(path)
            for i in ls:
                c_path = os.path.join(path, i)
                if os.path.isdir(c_path):
                    self.del_file(c_path)
                else:
                    os.remove(c_path)
            os.rmdir(path)

