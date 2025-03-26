#-*- coding:utf-8 -*-
# Create by  liuli
# Create on 2024/7/31
# import pandas as pd
import os
import sys
import openpyxl
import random
import time;
from datetime import datetime, timedelta
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

#pdf文件产生目录,结尾需要有斜杠
fileLocalPath = "D:\\temp\\"
tracepath = fileLocalPath + "debug.log"
reveal_item = "8月14日的披露事项"

# Excel文件路径
excel_file = 'E:\\code\\python\\buildpdf\\14三类.xlsx'

#产品代码所在的工作表的名字
sheetname = 'Sheet1'

#产品代码所在的列数，列从1开始计数
product_code_col = 1

#随机日期开始日期
start_date_random = datetime(2024, 1, 1)
#随机日期结束日期
end_date_random = datetime(2024, 7, 31)

#报告年度列表，程序会随机在这个列表中获取年份
report_years = [2024,2025,2026]

def savelog(strsen, display = True):
    if display:
        print(strsen)

    try:
        resultfile = open(tracepath,'a+')
        resultfile.write("[" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "]" + strsen + "\n")
        resultfile.close()
    except Exception as e:
        pass

# 生成一个包含给定文本的PDF文件。
# 将C:\Windows\Fonts\微软雅黑拷贝到你的python安装目录的 Lib\site-packages\reportlab\fonts 目录下
def generate_pdf(filename, text):
    """
    生成一个包含给定文本的PDF文件。
    """
    pdfmetrics.registerFont(TTFont('雅黑', 'msyhl.ttc'))
    c = canvas.Canvas(filename)
    c.setFont("雅黑", 12)
    text_width, text_height = c.stringWidth(text), c._fontsize
    text_height = text_height + 5
    x_center = (c._pagesize[0] - text_width) / 2
    y_start = c._pagesize[1] - text_height - 50  # 留出一些边距
    linemaxlen = 40
    if len(text) > linemaxlen:
        t = ""
        j = 0
        for i in text:
            t = t + i
            j = j + 1
            if j % linemaxlen == 0:
                c.drawString(20, y_start, t)
                t = ""
                y_start = y_start - text_height
        
        c.drawString(20, y_start, t)

        # savelog(t)
        # c.drawString(x_center, y_start, t)
    else:
        c.drawString(x_center, y_start, text)
    c.save()

def read_product_code(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.get_sheet_by_name(sheetname)
    result = set()
    for row in sheet.iter_rows(min_row=1, min_col=product_code_col, max_row=sheet.max_row, max_col=product_code_col, values_only=True):
        for cell in row:
            cells = str(cell)
            cstr = cells.strip()
            if len(cstr) < 11:
                continue
            else:
                result.add(cstr)
    return result

#获得文件数量标识
def get_file_num():
    k = random.randint(1,9)
    res = []
    for i in range(k):
        res.append(str(k) + "-" + str(i + 1))

    return res

#获得报告年度
def get_report_year():
    k = random.randint(1,len(report_years))
    return str(report_years[k - 1])

#获得披露类型
def get_reveal_type():
    k = random.randint(1,2)
    if k == 1:
        #定期披露
        return "R"
    else:
        #临时披露
        return "T"

#获得披露子类型
def get_reveal_sub_type(type):
    scheduled_lst = ["Q1","Q2","Q3","Q4","HY","YY","OR"]
    temporary_lst = ["IM","EN","OT"]
    if type == "R":
        k = random.randint(1,len(scheduled_lst))
        return scheduled_lst[k - 1]
    else:
        k = random.randint(1,len(temporary_lst))
        return temporary_lst[k - 1]

#获得报告类型
#披露类型为T时只可披露OR报告
def get_report_type(product_code,type,reveal_sub_type):
    if type == "R":
        if is_group_product(product_code):
            # 当前是组合类产品，组合类产品只可披露GL、TG、OR类型报告
            report_lst = ["GL","TG","OR"]
            k = random.randint(1,len(report_lst))
            return report_lst[k - 1]
        else:
            #当前是三类产品，三类产品不可披露GL类型报告
            report_lst = ["ST","TG","JD","NP","WP","OR"]
            k = random.randint(1,len(report_lst))
            #if reveal_sub_type == "OR" and report_lst[k - 1] == "OR":
               # k = random.randint(1,len(report_lst) - 1)
               # savelog("get_report_type reveal_sub_type and report_lst return OR,renew k:" + str(k))

            return report_lst[k - 1]
    else:
        return "OR"

# 三类产品以10开头
# 组合类产品以11为开头
# 当前产品代码是否是组合类产品
def is_group_product(product_code):
    p = str(product_code)
    if p[0] == '1' and p[1] == '1':
        return True
    else:
        return False

#获得随机日期
def get_random_date():
    days_diff = (end_date_random - start_date_random).days
    random_days = random.randint(0, days_diff)
    random_date = start_date_random + timedelta(days=random_days)
    return str(random_date.strftime('%Y%m%d'))

def buildpdf(excel_file,output_dir):
    product_code_set = read_product_code(excel_file)
    savelog("开始生成文件，产品代码共有" + str(len(product_code_set)) + "条数据。")
    id = 0
    for product_code in product_code_set:
        file_number_lst = get_file_num()
        savelog("开始为 " + str(product_code) + " 产品创建文件，大约要创建" + str(len(file_number_lst)) + "个文件。")
        report_year = get_report_year()
        reveal_type = get_reveal_type()
        reveal_sub_type = get_reveal_sub_type(reveal_type)
        report_type = get_report_type(product_code,reveal_type,reveal_sub_type)
        report_random_date = get_random_date()
        for file_identification in file_number_lst:
            #文件标识数
            filename = file_identification + "_" + report_year + "_" + str(product_code) + "_" + reveal_type + "_" + reveal_sub_type + "_"
            if reveal_type == "R":
                #定期披露，只有定期披露可以获得报告类型
                filename = filename + report_type
            else:
                #临时披露
                if reveal_sub_type == "EN":
                    #产品终止
                    filename = filename + report_random_date + reveal_item
                else:
                    filename = filename + reveal_item
            
            filename = filename + ".pdf"
            filepath = output_dir + filename
            savelog(filepath,False)
            id = id + 1
            generate_pdf(filepath,filename)

    savelog("共生成" + str(id) + "个文件。")

if __name__ == "__main__":
    output_dir = fileLocalPath # PDF文件存放的目录，确保末尾有斜杠
    if len(sys.argv) == 1:
        buildpdf(excel_file,output_dir)
    elif len(sys.argv) == 2:
        # test build pdf file function
        generate_pdf(sys.argv[1],reveal_item)
    elif len(sys.argv) == 3:
        buildpdf(sys.argv[1],sys.argv[2])
    else:
        print("wrong parameters!")