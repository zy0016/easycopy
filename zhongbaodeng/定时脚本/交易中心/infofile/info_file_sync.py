# coding=utf-8
# !/usr/bin/python
#-------------------------------------------------------------------------------
# Name:        模块1
# Purpose:
#
# Author:      lihuaizhong
#
# Created:     30-05-2019
# Copyright:   (c) lihuaizhong 2019
# Licence:     <your licence>
#-------------------------------------------------------------------------------
# 获取估值需要的产品信息列表和附件

import logging
import os
import os.path
import socket
import sys
import time

import paramiko
from openpyxl import Workbook
from SFTPSync import SFTPSync
try:
    import cx_Oracle
except ImportError as e:
    print("导入cx_Oracle模块失败,需要安装cx_Oracle\r\n{}",str(e))
    sys.exit(-1)

logging.basicConfig()
#oracle_base = 'E:\\app\\hspcadmin\\product'
#os.environ["ORACLE_BASE"] = oracle_base
#os.environ["ORACLE_HOME"] = oracle_base + '\\11.2.0\\client_1'
#os.environ["LD_LIBRARY_PATH"] = oracle_base + '\\11.2.0\\client_1\\LIB'
#os.environ["ORACLE_SID"] = 'kfzgdb'

# 附件catalog
catalog = 'TFUNDINFO'

class SSHSession(object):

    def __init__(self,fileSrcHost,fileSrcPort,fileSrcUserName,fileSrcPassword,dbip,dbport,dbSID,dbUserName,dbPassword,key_file=None):

        self.fileSrcHost = fileSrcHost
        self.fileSrcPort = fileSrcPort
        self.fileSrcUserName = fileSrcUserName
        self.fileSrcPassword = fileSrcPassword

        self.dbip = dbip
        self.dbport = dbport
        self.dbSID = dbSID
        self.dbusername = dbUserName
        self.dbpassword = dbPassword

        self.fileSrcSftp = self.getSFtp(self.fileSrcHost,self.fileSrcPort,self.fileSrcUserName,self.fileSrcPassword)
        self.cursor = self.getDbCursor(self.dbip,self.dbport,self.dbSID,self.dbusername,self.dbpassword)

    def getDbCursor(self,dbIp,dbPort,dbSID,dbUserName,dbPassword):
        try:
            dsn_tns = cx_Oracle.makedsn(dbIp, dbPort, service_name=dbSID)
            self.db = cx_Oracle.connect(dbUserName, dbPassword, dsn_tns)
            cursor = self.db.cursor()
            print("connet %s:%s/%s database successful!" % (dbIp,dbPort,dbSID))
            return cursor
        except cx_Oracle.Error as e:
            print("init connet database failed,exit!" + str(e))
            time.sleep(5)
            sys.exit(-1)

    def getSFtp(self, host, port ,userName,userPassword):
        print("connect getSFtp server %s:%s" % (host,port))
        sftp = SFTPSync(host,port)
        try:
            sftp.login(userName,userPassword)
            print("connect getSFtp server %s:%s Successful!!" % (host,port))
            return sftp
        except Exception as e:
            print("connect getSFtp server %s:%s failed,exit!" % (host,port))
            time.sleep(5)
            sys.exit(-1)

    def getFtp ( self ,host,port,userName,userPassword):
        print("connect fileserver %s:%s" % (host,port))
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(30 * 60 * 1000)
        sock.connect((host, port))
        t = paramiko.Transport(sock)
        t.start_client()
        # supposed to check for key in keys, but I don't much care right now to find the right notation
        if userPassword is not None:
            t.auth_password(userName, userPassword, fallback=False)
        else:
            raise Exception('Must supply either key_file or password')
        print("connect fileserver %s:%s Successful!!" % (host, port))
        return paramiko.SFTPClient.from_transport(t)

    def reconnectDB(self):
        try:
            print("reconnet %s:%s/%s database " % (self.dbip,self.dbport,self.dbSID))
            self.cursor = self.getDbCursor(self.dbip,self.dbport,self.dbSID,self.dbUserName,self.dbPassword)
            print("reconnet %s:%s/%s database successful!" % (self.dbip,self.dbport,self.dbSID))
        except cx_Oracle.Error as e:
            print("reconnet database failed,exit!" + str(e))
            time.sleep(5)
            sys.exit(-1)

    def disconnect(self):
        try:
            self.fileSrcSftp.close()
            self.db.close()
            print("Connection terminated.")
        except cx_Oracle.DatabaseError as e:
            print("close db  except." + str(e))

    # 获取产品代码并保存产品信息到excel
    def getFundcodes(self,templetId,startDate,endDate,filepath):
        fundcodes = set()
        sql = "select * from user_tables where table_name = 'TPTDALLINFO' "
        existData = self.cursor.execute(sql).fetchall()
        if(len(existData) > 0):
            sql = "drop table TPTDALLINFO "
            self.cursor.execute(sql)

        print('getFundcode.startDate = '+ startDate+',endDate='+endDate+',templetName='+ self.getTempleName(templetId))
        cursor = self.cursor.var(cx_Oracle.CURSOR)
        args = [startDate,endDate,'default',templetId,cursor]
        try:
            self.cursor.callproc('pkg_prdinfo.getbondbasedate', args)
            sql = "select * from TPTDALLINFO t order by t.c_fundcode"
            dataResult = self.cursor.execute(sql)
            # 将产品信息存放到excel里
            dataRecords = dataResult.fetchall()

            # 融资主体字段
            mandatories = ['g融资主体信息_融资主体','A类增信/收益保障_担保人全称','B类增信/收益保障_担保人全称']
            mandatoryList = list(mandatories)
            # 主体在原excel的列号
            mandatoryColList = list()
            # 主体在系统中的code
            mandatoryCodeList = list()

            if(len(dataRecords) > 1):
                extData = list()
                for col in range(len(dataRecords[0])):
                    if dataRecords[0][col] in mandatoryList:
                        mandatoryColList.append(col)
                        extData.append(dataRecords[0][col] + 'Code')
                mandatoryCodeList.append(extData)

                if len(mandatoryColList) > 0:
                    for i in range(1, len(dataRecords)):
                        extData = list()
                        for colNum in mandatoryColList:
                            mandatoryName = dataRecords[i][colNum]
                            extData.append(self.getMainCode(mandatoryName))
                        mandatoryCodeList.append(extData)

                result = list()
                for drRow in range(len(dataRecords)):
                    tmpRow = list()
                    for drCol in range(len(dataRecords[drRow])):
                        tmpRow.append(dataRecords[drRow][drCol])
                        if drCol in mandatoryColList:
                            tmpRow.append(mandatoryCodeList[drRow][mandatoryColList.index(drCol)])
                    result.append(tmpRow)

                self.write_to_excel(filepath, endDate, result, templetId)
                sql = "select c_fundcode,c_tafundcode  from tptdallinfo where c_fundcode != '-1'"
                dataresult = self.cursor.execute(sql)
            else:
                return fundcodes

            data = dataresult.fetchall()
            for r in range(len(data)):
                fundcode = data[r][0]
                tafundcode = data[r][1]
                fundcodes.add(fundcode+','+tafundcode)
            return fundcodes

        except Exception as e:
            print('获取templetId=' + templetId + '类型的产品信息异常'+str(e))
        return fundcodes

    # 获取融资主体/担保主体编码
    def getMainCode(self,name):
        if name is None:
            return ''
        if '' == name.strip():
            return name.strip()
        # sql = "  select C_ORGCODE from TPDTMANDATOR where c_name = '"+name+"'"
        sql = " select C_ORGCODE from TPDTMANDATOR where c_status not in ('0','2') and c_name ='"+name+"' order by length(c_orgcode) desc"
        dataresult = self.cursor.execute(sql)
        data = dataresult.fetchall()
        if len(data) > 0:
            return data[0][0]
        else:
            return ''

    # 保存数据到excel(不包含表头)
    def write_to_excel(self,path,date,data,templetId):
        print('保存产品"'+self.getTempleName(templetId)+'"详细信息到excel开始！')
        self.mkdir_ifnotexists(path+os.path.sep+date)
        os.chdir(path+os.path.sep+date)

        file_name = '{0}_{1}.xlsx'.format(date,self.getTempleName(templetId))
        wb = Workbook()
        sheet = wb.active
        sheet.title=date

        for r in range(len(data)):
            for c in range(len(data[r])):
                sheet.cell(row = r+1,column = c+1).value=data[r][c]
        wb.save(file_name)
        print('保存产品"'+self.getTempleName(templetId)+'"详细信息到excel结束！')

    # 保存数据到excel同时保存表头
    def write_to_excel_title(self,path,date,data,titles,templetId):
        print('保存产品详细信息到excel开始！')
        if len(data) > 0:
            self.mkdir_ifnotexists(path+os.path.sep+date)
            os.chdir(path+os.path.sep+date)
            file_name = '{0}_{1}.xlsx'.format(date,templetId)
            wb = Workbook()
            sheet = wb.active
            sheet.title=date

            for r in range(len(titles)):
                sheet.cell(row=1, column=r + 1).value = titles[r][0]

            for r in range(len(data)):
                for c in range(len(data[r])):
                    sheet.cell(row = r+2,column = c+1).value=data[r][c]
            wb.save(file_name)
        else:
            print('当天无‘'+templetId+'’产品数据！')
        print('保存产品详细信息到excel结束！')

    def getfilename(self,filename,filelocalpath):
        if os.path.exists(filelocalpath + os.path.sep + filename):
            id = 0
            while True:
                filenewname = ""
                id = id + 1
                extendid = filename.rindex('.')
                if extendid == -1:
                    print("附件没有扩展名!")
                    filenewname = filename + "(" + str(id) + ")"
                else:
                    filenewname = filename[0:extendid] + "(" + str(id) + ")" + filename[extendid:]
                
                if os.path.exists(filelocalpath + os.path.sep + filenewname):
                    continue
                else:
                    return filenewname
        else:
            return filename
    
    def savelog(self,strsen):
        print(strsen)
        try:
            resultfile = open("c:\\estimate\\infofile\\debug.log",'a+')
            resultfile.write("[" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "]" + strsen + "\n")
            resultfile.close()
        except Exception as e:
            pass

    def read_tblobstorage_table_attachmentCount(self,yesterday):
        sql_fileid = "select count(*) from interins.info_detail a, interins.info_file b,interins.file_info c where "
        sql_fileid = sql_fileid + "(a.INFO_ID=b.INFO_ID and b.FILE_ID=c.FILE_ID and (a.PRD_TYPE_2=1 or a.PRD_TYPE_2=3) and b.IS_DELETED=0 and a.INFO_STATUS=4 and a.INFO_TYPE=2 "
        sql_fileid = sql_fileid + "and a.INFO_START_DATE=" + str(yesterday) + ")"
        dataresult = self.cursor.execute(sql_fileid)
        data = dataresult.fetchall()
        return int(data[0][0])

    def read_tblobstorage_table_attachment(self,fileremotepath,filelocalpath,startDate,yesterday,pagecount):
        amount = self.read_tblobstorage_table_attachmentCount(yesterday)
        returnset = set()
        page_number = 0
        if amount % pagecount == 0:
            page_number = amount // pagecount
        else:
            page_number = amount // pagecount + 1

        sql = f"select * from (select ROWNUM r, iinfo.INFO_ID,iinfo.PRODUCT_CODE,iinfo.INFO_START_DATE,c.FILE_ID,c.FILE_PATH,b.FILE_NAME "
        sql = sql + "from interins.info_detail iinfo, interins.info_file b,interins.file_info c "
        sql = sql + "WHERE (iinfo.INFO_ID=b.INFO_ID and b.FILE_ID=c.FILE_ID and (iinfo.PRD_TYPE_2=1 or iinfo.PRD_TYPE_2=3) and b.IS_DELETED=0 "
        sql = sql + "and iinfo.INFO_STATUS=4 and iinfo.INFO_TYPE=2 and iinfo.INFO_START_DATE=" + str(yesterday)
        sql = sql + ") and ROWNUM <=:end_row) where r>= :start_row"
        pageid = 1
        start_row = 1
        end_row = pagecount
        while pageid <= page_number:
            dataresult = self.cursor.execute(sql, {"end_row": end_row,'start_row':start_row})
            rows = dataresult.fetchall()
            for datasingle in rows:
                self.savelog("filepath:" + str(datasingle[5]))
                localpath = self.getlocalpath(filelocalpath,str(datasingle[2]),startDate)
                localfile = self.getfilename(datasingle[6],localpath)
                remote_filename = str(datasingle[5])
                localdirpath = self.getFilesFromRemoteServer(filelocalpath,fileremotepath,str(datasingle[2]),"",remote_filename,localfile,startDate)
                returnset.add(localdirpath + "," + str(datasingle[2]))

            start_row = start_row + pagecount
            end_row = end_row + pagecount
            pageid = pageid + 1

        return returnset

    def getlocalpath(self,localpath,tafundcode,dataDate):
        return localpath +os.path.sep + dataDate + os.path.sep + 'infoannex'+ os.path.sep + tafundcode

    # 从远程文件服务器获取附件信息
    def getFilesFromRemoteServer(self,localpath,remotepath,tafundcode,prdname,refilename,lcfilename,dataDate):
        tmpprdpath = self.getlocalpath(localpath,tafundcode,dataDate)
        self.mkdir_ifnotexists(tmpprdpath)
        localtmpFile = tmpprdpath+os.path.sep+lcfilename
        remotefile=remotepath+'/'+refilename
        self.savelog('remotefile='+remotefile+'\t,localtmpFile='+localtmpFile)
        try:
            self.fileSrcSftp.get(remotefile, localtmpFile)
        except Exception as e:
            self.savelog('\t\t【获取附件异常,】' + str(e))
        
        return tmpprdpath
    
    # 获取附件信息
    def read_tblobstorage_table(self,fileremotepath,filelocalpath,fundcode,startDate,tafundcode):
        print('【获取产品'+fundcode+'_'+tafundcode+'附件信息开始】')
        named_params = { 'subkeyid':fundcode}
        # sql = "SELECT  a.c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename FROM tblobstorage bs,tfundinfo a  WHERE bs.c_subkeyid=a.c_fundcode and c_catalog=:ccatalog AND c_subkeyid=:subkeyid"
        sql = " SELECT  a.c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename " \
              " FROM tblobstorage bs,tfundinfo a  " \
              " WHERE bs.c_subkeyid=a.c_fundcode " \
              " and (instr(bs.c_filename,'募集') > 0 or instr(bs.c_filename,'合同') > 0 or instr(bs.c_filename,'风险评估') > 0 " \
              "      or instr(bs.c_filename,'评级') > 0 or instr(bs.c_filename,'认购协议') > 0 or instr(bs.c_filename,'标准条款') > 0 or instr(bs.c_filename,'托管合同') > 0 or instr(bs.c_filename,'托管协议') > 0 or instr(bs.c_filename,'主定义表') > 0 ) " \
              " and (c_catalog='TFUNDINFO' or c_catalog='TFUNDINFOT' or c_catalog='TFUNDINFOT_REGISTER')  " \
              " AND c_subkeyid=:subkeyid "
        dataresult = self.cursor.execute(sql, named_params)
        data = dataresult.fetchall()
        for r in range(len(data)):
            print("\ttafundcode=%s, prdName=%s ,remotePath=%s ,localFileName=%s"%(tafundcode,data[r][0],data[r][1],data[r][2]) )
            self.getFileFromRemoteServer(filelocalpath,fileremotepath,tafundcode,data[r][0],data[r][1],data[r][2],startDate)
        print('【产品'+fundcode+'_'+tafundcode+'共获取'+str(len(data))+'条附件记录】')

    # 从远程文件服务器获取附件信息
    def getFileFromRemoteServer(self,localpath,remotepath,tafundcode,prdname,refilename,lcfilename,dataDate):
        tmpprdpath = localpath +os.path.sep + dataDate + os.path.sep + 'annex'+ os.path.sep + tafundcode
        self.mkdir_ifnotexists(tmpprdpath)
        localtmpFile = tmpprdpath+os.path.sep+lcfilename
        remotefile=remotepath+'/'+refilename
        print('\t\tremotefile='+remotefile+'\t,localtmpFile='+localtmpFile)
        try:
            self.fileSrcSftp.get(remotefile, localtmpFile)
        # shutil.copy(localfile,localtmpFile)
        except Exception as e:
            print('\t\t【获取附件异常,】' + str(e))

    # 获取存续期兑付附件信息
    def read_tblobstorage_table_Cash(self,fileremotepath,filelocalpath,lid,startDate,tafundcode):
        print('【获取产品'+tafundcode+'附件信息开始】')
        named_params = { 'csubkeyid':lid}
        # sql = "SELECT  a.c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename FROM tblobstorage bs,tfundinfo a  WHERE bs.c_subkeyid=a.c_fundcode and c_catalog=:ccatalog AND c_subkeyid=:subkeyid"
        sql = " SELECT  'nan' c_fundname,trunc(bs.l_storageid/10,0)||'/'||bs.l_storageid||'baccessory' as remotefile,trim(bs.c_filename) as c_filename " \
              " FROM tblobstorage bs   " \
              " WHERE (bs.c_catalog='TPDTDURATIONPROGRAM' )  " \
              " and c_subkeyid=:csubkeyid "
        print('sql2='+sql)
        dataresult = self.cursor.execute(sql, named_params)
        data = dataresult.fetchall()
        for r in range(len(data)):
            print("\ttafundcode=%s, prdName=%s ,remotePath=%s ,localFileName=%s"%(tafundcode,data[r][0],data[r][1],data[r][2]) )
            self.getFileFromRemoteServer_Cash(filelocalpath,fileremotepath,tafundcode,data[r][0],data[r][1],data[r][2],startDate)
        print('【产品'+tafundcode+'共获取'+str(len(data))+'条附件记录】')

    # 从远程文件服务器获取附件信息
    def getFileFromRemoteServer_Cash(self,localpath,remotepath,tafundcode,prdname,refilename,lcfilename,dataDate):
        tmpprdpath = localpath +os.path.sep + dataDate + os.path.sep + 'annex-cash'+ os.path.sep + tafundcode
        self.mkdir_ifnotexists(tmpprdpath)
        localtmpFile = tmpprdpath+os.path.sep+lcfilename
        remotefile=remotepath+'/'+refilename
        print('\t\tremotefile='+remotefile+'\t,localtmpFile='+localtmpFile)
        try:
            self.fileSrcSftp.get(remotefile, localtmpFile)
        # shutil.copy(localfile,localtmpFile)
        except Exception as e:
            print('\t\t【获取附件异常,】' + str(e))

    # 创建目录
    def mkdir_ifnotexists(self, path):
        path=path.strip()
        path=path.rstrip("\\")
        isExists=os.path.exists(path)

        # 判断结果
        if not isExists:
            # 如果不存在则创建目录
            os.makedirs(path)

            print('\t'+path+' 创建成功')
            return True
        else:
            # 如果目录存在则不创建，并提示目录已存在
            # print(path+' 目录已存在')
            return False

    # 从产品中心获取存续期产品
    def getDurationFundinfo(self,startDate,endDate,filelocalpath):
        print('从产品中心获取存续期startDate='+startDate+'，endDate='+endDate+'数据开始')
        sql="select  t.l_id 序号,  a.c_fundcode   产品代码, a.c_tafundcode TA产品代码, a.c_fundname 产品名称, decode( a.L_PDTTEMPLETID,'6','债权投资计划','8','资产支持计划') 产品类型," \
            " decode( t.c_type,'0','收益分配','1','赎回','2','回售','3','分期还本','4','兑付','5','清算','6','分红') 业务类型,  c_date 业务发生日期 ,  f_totalamount 业务发生总额, s_severaltime 期次," \
            " decode( c_paymenttype,'1','正常到期','2','提前还款')  兑付类型,  f_totalshare 业务发生总份数, t.d_createdate 创建日期,  f_afterfacevalue 执行后份额面值,  t.c_memo 备注" \
            " FROM tpdtdurationprogram t, tfundinfo a, member.operators o" \
            " WHERE t.c_fundcode=a.c_fundcode and  (a.L_PDTTEMPLETID in('6','8'))" \
            "   and t.c_creator = o.oper_id(+)" \
            "   and t.c_type='4'" \
            " and to_char(t.l_id) in (select biz_flow_uuid_ from jbpm4_ext_htask_ j where j.PROC_NAME_ like  '%存续期方案流程%'"\
            " and j.TASKOUTCOMEACTION = '复审通过' and j.activity_name_ = '存续期方案复审' "\
            " and to_char(j.END_, 'yyyyMMdd') >= '"+startDate+"' and to_char(j.END_, 'yyyyMMdd') <= '"+endDate+"' )"\
            " ORDER BY  t.c_type asc, t.d_createdate desc"
        dataResult = self.cursor.execute(sql)
        # 将产品信息存放到excel里
        dataRecords = dataResult.fetchall()
        self.write_to_excel_title(filelocalpath,endDate,dataRecords,dataResult.description,'存续期-兑付')
        print('从产品中心获取存续期startDate=' + startDate + '，endDate=' + endDate + '数据结束')

    # 从产品中心获取当日存续期兑付产品编码
    def getDurationFundinfo_Cash(self,templetId,startDate,endDate,filelocalpath):
        print('从产品中心获取存续期startDate='+startDate+'，endDate='+endDate+'数据开始')
        sql="select  to_char(t.l_id) as l_id,a.c_tafundcode "\
            " FROM tpdtdurationprogram t, tfundinfo a, member.operators o" \
            " WHERE t.c_fundcode=a.c_fundcode and a.L_PDTTEMPLETID= '"+templetId+"' " \
            "   and t.c_creator = o.oper_id(+)" \
            "   and t.c_type='4' " \
            " and to_char(t.l_id) in (select biz_flow_uuid_ from jbpm4_ext_htask_ j where j.PROC_NAME_ like  '%存续期方案流程%'"\
            " and j.TASKOUTCOMEACTION = '复审通过' and j.activity_name_ = '存续期方案复审' "\
            " and to_char(j.END_, 'yyyyMMdd') >= '"+startDate+"' and to_char(j.END_, 'yyyyMMdd') <= '"+endDate+"' )"\
            " ORDER BY  t.c_type asc, t.d_createdate desc"
        fundcodes=set()
        print('sql=' + sql)
        try:
            dataResult = self.cursor.execute(sql)
            # 将产品信息存放到excel里
            dataRecords = dataResult.fetchall()
            for r in range(len(dataRecords)):
                lid = dataRecords[r][0]
                ctafundcode = dataRecords[r][1]
                fundcodes.add(lid+','+ctafundcode)
            return fundcodes
            print('从产品中心获取存续期兑付startDate=' + startDate + '，endDate=' + endDate + '数据结束')
        except Exception as e:
            print('获取templetId=' + templetId + '类型的产品信息兑付异常'+str(e))
        return fundcodes

    # 从ta获取存续期产品信息
    def getFundInfoByTA(self,filelocalpath,startDate,endDate):
        print('从TA获取存续期startDate=' + startDate + '，endDate=' + endDate + '数据开始')
        sql = "SELECT a.c_fundcode 产品代码," \
                      " b.c_bondfullname 产品名称," \
                      " decode(b.c_producttype, '1', '债权', '2', '股权', '3', '资产支持计划') 产品类型," \
                      " decode(c_rdmflag, 'l', '回售', 'm', '赎回', 'i', '分期还本') 业务类型," \
                      " a.d_rdmdate 业务发生日期," \
                      " a.d_impdate 导入日期," \
                      " b.f_issueprice 执行后份额面值," \
                      " a.sumbalance 业务发生总额," \
                      " a.sumshare 业务发生总份数" \
            " FROM (SELECT a.c_fundcode, a.c_rdmflag, a.d_rdmdate, a.d_impdate," \
                      "    sum(a.f_redeembalance) sumbalance, sum(a.f_redeemshares) sumshare" \
                    " FROM tbatchrightline a" \
                    " group by a.c_fundcode, a.c_rdmflag, a.d_rdmdate, a.d_impdate) a, tbondinfo b" \
            " where a.c_fundcode = b.c_bondcode " \
              " and a.d_impdate >= '"+startDate+"' and a.d_impdate <= '"+endDate+"'" \
            " union all" \
            " SELECT a.c_fundcode 产品代码," \
                    " b.c_bondfullname 产品名称," \
                    " decode(b.c_producttype, '1', '债权', '2', '股权', '3', '资产支持计划') 产品类型," \
                    " '收益分配' 业务类型," \
                    " a.d_regdate 业务发生日期," \
                    " a.d_cdate 导入日期," \
                    " b.f_issueprice 执行后份额面值," \
                    " a.sumbalance 业务发生总额," \
                    " a.sumshare 业务发生总份数" \
            " FROM (SELECT a.c_fundcode, a.d_regdate, a.d_cdate," \
                    " sum(a.f_realbalance) sumbalance, sum(a.f_realshares) sumshare" \
                    " FROM tdividenddetail a" \
                    " group by a.c_fundcode, a.d_regdate, a.d_cdate) a,tbondinfo b" \
            " where a.c_fundcode = b.c_bondcode " \
                    " and a.d_cdate >= '"+startDate+"' and a.d_cdate <= '"+endDate+"'" \

        dataResult = self.cursor.execute(sql)
        # 将产品信息存放到excel里
        dataRecords = dataResult.fetchall()
        self.write_to_excel_title(filelocalpath, endDate, dataRecords, dataResult.description, '存续期-无兑付')
        print('从TA获取存续期startDate=' + startDate + '，endDate=' + endDate + '数据结束')

    def getTempleName(self,templetId):
        if templetId == '8':
            return '资产支持计划'
        elif templetId == '6':
            return '债权'
        elif templetId == '7':
            return '股权'
        else:
            return templetId

    def sendToDestFtp(self,localDir,date,fileDestHost,fileDestPort,fileDestUsername,fileDestPassword,fileDestRemotePath):
        fileDestSftp = self.getFtp(fileDestHost,fileDestPort,fileDestUsername,fileDestPassword)

        fileDestSftp.stat(fileDestRemotePath)
        stdin, stdout, stderr = fileDestSftp.exec_command('ls DIR')
        if stdout.readline() != '':
            print("exist")
        else:
            print("not exist")

        fileDestSftp.chdir(fileDestRemotePath)
        fileDestSftp.mkdir(date);
        fileDestSftp.put()

    def get_filelist(self,dir, Filelist):
        newDir = dir
        if os.path.isfile(dir):
            Filelist.append(dir)
        elif os.path.isdir(dir):
            for s in os.listdir(dir):
                newDir = os.path.join(dir, s)
                self.get_filelist(newDir, Filelist)
        return Filelist

    # 删除
    def del_file(self,path):
        if os.path.exists(path):
            ls = os.listdir(path)
            for i in ls:
                c_path = os.path.join(path, i)
                if os.path.isdir(c_path):
                    self.del_file(c_path)
                else:
                    os.remove(c_path)
            os.rmdir(path)

